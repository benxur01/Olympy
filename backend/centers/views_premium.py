"""Tashkilotlar (markazlar) uchun premium analitika va hisobot endpoint'lari.

Barchasi `/api/centers/<id>/...` ostida mount qilinadi (centers/urls.py).
Ruxsat: `user_can_manage_center` (owner/manager + platform admin). Ba'zilari
qo'shimcha `center.is_premium` tekshiruvini talab qiladi (T1, T4).

Mavjud `centers/views.py` kattalashib ketmasligi uchun yangi premium
endpoint'lar shu faylga ajratildi — eski endpoint'lar buzilmaydi.
"""
import io
from datetime import timedelta

from django.db.models import Avg, Count, Max, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from rest_framework import status as http_status
from rest_framework.decorators import (
    api_view, permission_classes, renderer_classes,
)
from rest_framework.permissions import IsAuthenticated
from rest_framework.renderers import BaseRenderer, JSONRenderer
from rest_framework.response import Response

from .models import CenterMembership, EducationCenter
from .services import user_can_manage_center


# DRF `URL_FORMAT_OVERRIDE='format'` — ya'ni `?format=xlsx` so'rovi content
# negotiation tomonidan tutiladi va mos renderer topilmasa 404 beradi. T6
# eksport endpoint'i `?format=xlsx|csv` ni qabul qilishi uchun shu formatlarni
# taniydigan passthrough renderer'lar ro'yxatdan o'tkaziladi (faqat shu view'da).
class _PassthroughXlsxRenderer(BaseRenderer):
    media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    format = 'xlsx'

    def render(self, data, accepted_media_type=None, renderer_context=None):
        return data


class _PassthroughCsvRenderer(BaseRenderer):
    media_type = 'text/csv'
    format = 'csv'

    def render(self, data, accepted_media_type=None, renderer_context=None):
        return data


def _premium_center_required():
    return Response(
        {'detail': "Bu funksiya premium markazlar uchun.", 'upgrade_required': True},
        status=http_status.HTTP_403_FORBIDDEN,
    )


def _period_start(period):
    """`period` (week|month|all) uchun boshlang'ich sanani qaytaradi.

    None — cheklovsiz (barcha vaqt). Aks holda timezone-aware datetime.
    """
    now = timezone.now()
    if period == 'week':
        return now - timedelta(days=7)
    if period == 'month':
        return now - timedelta(days=30)
    return None


def _mask_phone(phone):
    """Telefon raqamni qisman yashiradi: +998 90 *** ** 67 ko'rinishida."""
    if not phone:
        return '—'
    digits = ''.join(ch for ch in str(phone) if ch.isdigit())
    if len(digits) < 7:
        return '***'
    return f"{digits[:5]}***{digits[-2:]}"


def _student_user_ids(center):
    """Markazning tasdiqlangan o'quvchi user_id'lari ro'yxati."""
    return list(
        CenterMembership.objects.filter(
            center=center,
            role=CenterMembership.ROLE_STUDENT,
            status=CenterMembership.STATUS_APPROVED,
        ).values_list('user_id', flat=True)
    )


# ─── T1. O'quvchi taqqoslash jadvali ────────────────────────────────────────


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def member_comparison(request, center_id):
    """GET /api/centers/<id>/member-comparison/?tag=&subject=&period=

    Markaz o'quvchilarining shu markaz olimpiadalaridagi natijalari bo'yicha
    taqqoslash jadvali. Premium markaz uchun. Filtrlar:
      - tag: group_tag bo'yicha (faqat shu guruh o'quvchilari)
      - subject: olimpiada fani bo'yicha
      - period: week|month|all (default all)
    Javob: [{full_name, total_score, attempt_count, avg_score, rank}]
    """
    from attempts.models import TestAttempt
    from django.db.models import Sum

    center = get_object_or_404(EducationCenter, pk=center_id)
    if not user_can_manage_center(request.user, center):
        return Response({'detail': 'Forbidden'}, status=http_status.HTTP_403_FORBIDDEN)
    if not center.is_premium:
        return _premium_center_required()

    tag = (request.query_params.get('tag') or '').strip()
    subject = (request.query_params.get('subject') or '').strip()
    period = (request.query_params.get('period') or 'all').strip()

    members_qs = CenterMembership.objects.filter(
        center=center,
        role=CenterMembership.ROLE_STUDENT,
        status=CenterMembership.STATUS_APPROVED,
    ).select_related('user')
    if tag:
        members_qs = members_qs.filter(group_tag=tag)
    members = list(members_qs)
    user_ids = [m.user_id for m in members]
    if not user_ids:
        return Response([])

    attempts_qs = TestAttempt.objects.filter(
        user_id__in=user_ids,
        olympiad__center=center,
        olympiad__is_deleted=False,
        disqualified=False,
    )
    if subject:
        attempts_qs = attempts_qs.filter(olympiad__subject__iexact=subject)
    start = _period_start(period)
    if start is not None:
        attempts_qs = attempts_qs.filter(submitted_at__gte=start)

    # Bitta GROUP BY so'rov bilan har o'quvchining agregatlari.
    stats_map = {
        row['user_id']: row
        for row in attempts_qs.values('user_id').annotate(
            total_score=Sum('score'),
            attempt_count=Count('id'),
            avg_score=Avg('score'),
        )
    }

    rows = []
    for m in members:
        st = stats_map.get(m.user_id)
        rows.append({
            'user_id': m.user_id,
            'full_name': m.user.full_name or m.user.normalized_phone or '—',
            'group_tag': m.group_tag or '',
            'total_score': (st['total_score'] if st else 0) or 0,
            'attempt_count': (st['attempt_count'] if st else 0) or 0,
            'avg_score': round((st['avg_score'] if st else 0) or 0, 1),
        })
    # Reyting: o'rtacha ball desc, keyin urinishlar soni desc tie-break.
    rows.sort(key=lambda r: (-r['avg_score'], -r['attempt_count']))
    for i, r in enumerate(rows, start=1):
        r['rank'] = i
    return Response(rows)


# ─── T2. Haftalik/oylik hisobot (PDF / JSON) ────────────────────────────────


def _build_center_report_data(center, period):
    """Markaz hisoboti uchun agregat ma'lumotlarni yig'adi (T2 uchun umumiy)."""
    from attempts.models import TestAttempt
    from olympiads.models import Olympiad

    start = _period_start(period)
    student_count = CenterMembership.objects.filter(
        center=center,
        role=CenterMembership.ROLE_STUDENT,
        status=CenterMembership.STATUS_APPROVED,
    ).count()

    olympiads_qs = Olympiad.objects.filter(center=center, is_deleted=False)
    attempts_qs = TestAttempt.objects.filter(
        olympiad__center=center, olympiad__is_deleted=False, disqualified=False,
    )
    if start is not None:
        olympiads_qs = olympiads_qs.filter(created_at__gte=start)
        attempts_qs = attempts_qs.filter(submitted_at__gte=start)

    agg = attempts_qs.aggregate(avg=Avg('score'), total=Count('id'))
    top_rows = (
        attempts_qs.values('user_id')
        .annotate(avg_score=Avg('score'), attempts=Count('id'))
        .order_by('-avg_score', '-attempts')[:5]
    )
    top_rows = list(top_rows)
    from accounts.models import User
    name_map = {
        u.id: (u.full_name or u.normalized_phone or '—')
        for u in User.objects.filter(id__in=[r['user_id'] for r in top_rows])
    }
    top5 = [
        {
            'rank': i,
            'full_name': name_map.get(r['user_id'], '—'),
            'avg_score': round(r['avg_score'] or 0, 1),
            'attempts': r['attempts'] or 0,
        }
        for i, r in enumerate(top_rows, start=1)
    ]
    period_label = {'week': 'Haftalik', 'month': 'Oylik'}.get(period, 'Umumiy')
    return {
        'center_name': center.name,
        'period': period,
        'period_label': period_label,
        'date': timezone.now().strftime('%Y-%m-%d'),
        'students_count': student_count,
        'olympiads_count': olympiads_qs.count(),
        'average_score': round(agg['avg'] or 0, 1),
        'total_attempts': agg['total'] or 0,
        'top_students': top5,
    }


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def report_json(request, center_id):
    """GET /api/centers/<id>/report-json/?period=week|month — JSON hisobot."""
    center = get_object_or_404(EducationCenter, pk=center_id)
    if not user_can_manage_center(request.user, center):
        return Response({'detail': 'Forbidden'}, status=http_status.HTTP_403_FORBIDDEN)
    if not center.is_premium:
        return _premium_center_required()
    period = (request.query_params.get('period') or 'week').strip()
    return Response(_build_center_report_data(center, period))


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def report_pdf(request, center_id):
    """GET /api/centers/<id>/report-pdf/?period=week|month — PDF hisobot.

    reportlab/weasyprint o'rnatilmagan (loyhada Pillow bor) — shu sababli
    Pillow orqali rasm chiziladi va PDF formatda saqlanadi (accounts/reports.py
    bilan bir xil yondashuv). Pillow ham mavjud bo'lmasa JSON qaytaramiz.
    """
    center = get_object_or_404(EducationCenter, pk=center_id)
    if not user_can_manage_center(request.user, center):
        return Response({'detail': 'Forbidden'}, status=http_status.HTTP_403_FORBIDDEN)
    if not center.is_premium:
        return _premium_center_required()
    period = (request.query_params.get('period') or 'week').strip()
    data = _build_center_report_data(center, period)

    try:
        pdf_bytes = _render_center_report_pdf(data)
    except Exception:
        import logging
        logging.getLogger(__name__).exception('center report pdf render failed')
        # Fallback: PDF chizilmasa JSON qaytaramiz (spec talabiga ko'ra).
        return Response(data)

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    safe_name = ''.join(
        ch for ch in (center.name or 'markaz')
        if ch.isalnum() or ch in (' ', '_', '-')
    )[:50].strip().replace(' ', '_') or 'markaz'
    response['Content-Disposition'] = (
        f'attachment; filename="olympy-hisobot-{safe_name}-{period}.pdf"'
    )
    return response


def _render_center_report_pdf(data):
    """Markaz hisobotini Pillow orqali A4 rasm sifatida chizib PDF qaytaradi."""
    from PIL import Image, ImageDraw

    from accounts.reports import _load_font, _measure

    width, height = 1200, 1700
    bg_color = (6, 8, 24)
    card_bg = (14, 17, 45)
    border_color = (32, 41, 95)
    accent = (99, 102, 241)
    text_white = (255, 255, 255)
    text_gray = (148, 163, 184)
    text_muted = (100, 116, 139)

    img = Image.new('RGB', (width, height), bg_color)
    draw = ImageDraw.Draw(img)

    font_logo = _load_font(28, bold=True)
    font_title = _load_font(34, bold=True)
    font_subtitle = _load_font(20)
    font_section = _load_font(26, bold=True)
    font_body = _load_font(18)
    font_body_bold = _load_font(18, bold=True)
    font_metric_num = _load_font(38, bold=True)
    font_metric_lbl = _load_font(14)

    # Header
    draw.rounded_rectangle([80, 80, 130, 130], radius=12, fill=accent)
    w_o, h_o = _measure(draw, 'O', font_logo)
    draw.text((105 - w_o / 2, 105 - h_o / 2), 'O', fill=text_white, font=font_logo)
    draw.text((145, 90), 'OLYMPY', fill=text_white, font=font_logo)
    draw.text((80, 160), f"{data['period_label'].upper()} MARKAZ HISOBOTI", fill=text_white, font=font_title)
    draw.text((80, 215), f"Sana: {data['date']}", fill=text_gray, font=font_subtitle)
    draw.line([80, 255, 1120, 255], fill=border_color, width=2)

    # Center name card
    draw.rounded_rectangle([80, 280, 1120, 400], radius=24, fill=card_bg, outline=border_color, width=2)
    draw.text((110, 310), 'Tashkilot', fill=text_gray, font=font_body)
    draw.text((110, 340), data['center_name'][:48], fill=text_white, font=font_section)

    # Metric cards (3 grid)
    metrics = [
        {'lbl': "O'QUVCHILAR", 'val': f"{data['students_count']} ta", 'col': (99, 102, 241)},
        {'lbl': 'OLIMPIADALAR', 'val': f"{data['olympiads_count']} ta", 'col': (14, 165, 233)},
        {'lbl': "O'RTACHA BALL", 'val': f"{data['average_score']}%", 'col': (16, 185, 129)},
    ]
    card_w = 320
    card_h = 150
    gap = 40
    for i, m in enumerate(metrics):
        x1 = 80 + i * (card_w + gap)
        y1 = 440
        x2 = x1 + card_w
        y2 = y1 + card_h
        draw.rounded_rectangle([x1, y1, x2, y2], radius=20, fill=card_bg, outline=border_color, width=2)
        draw.line([x1 + 25, y1 + 12, x2 - 25, y1 + 12], fill=m['col'], width=4)
        w_val, _ = _measure(draw, m['val'], font_metric_num)
        draw.text((x1 + (card_w - w_val) / 2, y1 + 45), m['val'], fill=text_white, font=font_metric_num)
        w_lbl, _ = _measure(draw, m['lbl'], font_metric_lbl)
        draw.text((x1 + (card_w - w_lbl) / 2, y1 + 105), m['lbl'], fill=text_gray, font=font_metric_lbl)

    draw.text((80, 640), f"Jami urinishlar: {data['total_attempts']} ta", fill=text_gray, font=font_subtitle)

    # Top 5 students
    draw.text((80, 720), 'TOP 5 OʻQUVCHI', fill=text_white, font=font_section)
    box_y1 = 770
    box_h = 460
    draw.rounded_rectangle([80, box_y1, 1120, box_y1 + box_h], radius=24, fill=card_bg, outline=border_color, width=2)
    top = data['top_students']
    if not top:
        empty = "Hozircha natijalar mavjud emas."
        w_e, _ = _measure(draw, empty, font_subtitle)
        draw.text((600 - w_e / 2, box_y1 + box_h / 2 - 10), empty, fill=text_gray, font=font_subtitle)
    else:
        for idx, s in enumerate(top):
            row_y = box_y1 + 45 + idx * 80
            draw.text((120, row_y), f"{s['rank']}. {s['full_name'][:30]}", fill=text_white, font=font_body_bold)
            draw.text((120, row_y + 28), f"Urinishlar: {s['attempts']} ta", fill=text_gray, font=font_metric_lbl)
            pct_str = f"{s['avg_score']}%"
            w_p, _ = _measure(draw, pct_str, font_section)
            draw.text((1080 - w_p, row_y + 4), pct_str, fill=accent, font=font_section)

    # Footer
    draw.text((80, 1600), 'Ushbu hisobot Olympy tahlil tizimi tomonidan avtomatik shakllantirildi.', fill=text_muted, font=font_subtitle)

    buf = io.BytesIO()
    img.save(buf, format='PDF')
    buf.seek(0)
    return buf.getvalue()


# ─── T3. Faollik ogohlantirish (inactive students) ──────────────────────────


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def inactive_students(request, center_id):
    """GET /api/centers/<id>/inactive-students/?days=14

    Oxirgi `days` kun ichida faol bo'lmagan (last_active_date eski yoki yo'q)
    o'quvchilar ro'yxati. Faollik `User.last_active_date` (streak yangilanish
    sanasi) bo'yicha aniqlanadi.
    Javob: [{full_name, phone (masklangan), last_active, days_inactive}]
    """
    center = get_object_or_404(EducationCenter, pk=center_id)
    if not user_can_manage_center(request.user, center):
        return Response({'detail': 'Forbidden'}, status=http_status.HTTP_403_FORBIDDEN)
    if not center.is_premium:
        return _premium_center_required()

    try:
        days = int(request.query_params.get('days') or 14)
    except (TypeError, ValueError):
        days = 14
    days = max(1, min(days, 365))

    today = timezone.now().date()
    cutoff = today - timedelta(days=days)

    # Nofaollik filtri ORM darajasida bajariladi (Python loop'da emas):
    # last_active_date NULL (hech qachon faol bo'lmagan) yoki cutoff'dan oldin
    # bo'lgan o'quvchilar. Bu butun a'zolar ro'yxatini xotiraga yuklab keyin
    # filtrlash o'rniga DB'da WHERE bilan kerakli qatorlarni qaytaradi.
    members = list(
        CenterMembership.objects.filter(
            center=center,
            role=CenterMembership.ROLE_STUDENT,
            status=CenterMembership.STATUS_APPROVED,
        ).filter(
            Q(user__last_active_date__isnull=True)
            | Q(user__last_active_date__lte=cutoff)
        ).select_related('user')
    )
    rows = []
    for m in members:
        u = m.user
        last = u.last_active_date
        if last is not None:
            days_inactive = (today - last).days
            last_str = last.isoformat()
        else:
            # Hech qachon faol bo'lmagan — hisob yaratilganidan beri.
            days_inactive = None
            last_str = None
        rows.append({
            'user_id': u.id,
            'full_name': u.full_name or '—',
            'phone': _mask_phone(u.normalized_phone or u.phone),
            'group_tag': m.group_tag or '',
            'last_active': last_str,
            'days_inactive': days_inactive,
        })
    # Eng uzoq nofaol birinchi (None = hech qachon faol bo'lmagan, eng tepada).
    rows.sort(key=lambda r: (r['days_inactive'] is not None, r['days_inactive'] or 0), reverse=True)
    return Response(rows)


# ─── T4. Savollar banki analitikasi ─────────────────────────────────────────


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def question_analytics(request, center_id):
    """GET /api/centers/<id>/question-analytics/?subject=

    Markaz savollarining qiyinlik analitikasi: har savol uchun nechta marta
    javob berilgan va necha foiz xato. `wrong_percentage` bo'yicha desc
    tartiblangan (eng qiyin savollar birinchi). Premium markaz uchun.
    """
    from questions.models import Question
    from attempts.models import TestAttempt

    center = get_object_or_404(EducationCenter, pk=center_id)
    if not user_can_manage_center(request.user, center):
        return Response({'detail': 'Forbidden'}, status=http_status.HTTP_403_FORBIDDEN)
    if not center.is_premium:
        return _premium_center_required()

    subject = (request.query_params.get('subject') or '').strip()
    questions_qs = Question.objects.filter(center=center)
    if subject:
        questions_qs = questions_qs.filter(subject__iexact=subject)
    questions = list(questions_qs.only('id', 'text', 'correct_answer', 'subject'))
    if not questions:
        return Response([])
    qmap = {q.id: q for q in questions}
    total_attempts_map = {q.id: 0 for q in questions}
    wrong_map = {q.id: 0 for q in questions}

    # Markaz olimpiadalaridagi barcha valid attempts'lar answers JSON'ini
    # oqim ko'rinishida o'qiymiz (xotirani tejash uchun).
    answers_stream = (
        TestAttempt.objects
        .filter(olympiad__center=center, disqualified=False)
        .values_list('answers', flat=True)
        .iterator(chunk_size=500)
    )
    for ans in answers_stream:
        if not isinstance(ans, dict):
            continue
        for k, v in ans.items():
            try:
                qid = int(k)
            except (TypeError, ValueError):
                continue
            q = qmap.get(qid)
            if not q:
                continue
            total_attempts_map[qid] += 1
            try:
                if int(v) != q.correct_answer:
                    wrong_map[qid] += 1
            except (TypeError, ValueError):
                wrong_map[qid] += 1

    rows = []
    for q in questions:
        ta = total_attempts_map.get(q.id, 0)
        wc = wrong_map.get(q.id, 0)
        wrong_pct = round((wc / ta) * 100, 1) if ta else 0.0
        rows.append({
            'question_id': q.id,
            'text': (q.text or '')[:120],
            'subject': q.subject or '',
            'total_attempts': ta,
            'wrong_count': wc,
            'wrong_percentage': wrong_pct,
        })
    # Eng qiyin (eng ko'p xato) birinchi; teng bo'lsa ko'proq urinilgani.
    rows.sort(key=lambda r: (-r['wrong_percentage'], -r['total_attempts']))
    return Response(rows)


# ─── T5. Guruh teglari bo'yicha taqqoslash ──────────────────────────────────


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def tag_comparison(request, center_id):
    """GET /api/centers/<id>/tag-comparison/

    Har bir `group_tag` uchun: o'quvchilar soni, o'rtacha ball, jami urinishlar.
    Teg belgilanmagan o'quvchilar "(teg yo'q)" guruhiga yig'iladi.
    Javob: [{tag, student_count, avg_score, total_attempts}]
    """
    from attempts.models import TestAttempt
    from django.db.models import Sum

    center = get_object_or_404(EducationCenter, pk=center_id)
    if not user_can_manage_center(request.user, center):
        return Response({'detail': 'Forbidden'}, status=http_status.HTTP_403_FORBIDDEN)
    if not center.is_premium:
        return _premium_center_required()

    members = list(
        CenterMembership.objects.filter(
            center=center,
            role=CenterMembership.ROLE_STUDENT,
            status=CenterMembership.STATUS_APPROVED,
        ).values('user_id', 'group_tag')
    )
    if not members:
        return Response([])

    # user_id → tag map; tag → o'quvchilar soni.
    tag_by_user = {}
    tag_counts = {}
    for m in members:
        tag = (m['group_tag'] or '').strip() or '(teg yo\'q)'
        tag_by_user[m['user_id']] = tag
        tag_counts[tag] = tag_counts.get(tag, 0) + 1

    user_ids = list(tag_by_user.keys())
    # Har o'quvchining markazdagi agregati (bitta GROUP BY).
    per_user = {
        row['user_id']: row
        for row in TestAttempt.objects.filter(
            user_id__in=user_ids,
            olympiad__center=center,
            olympiad__is_deleted=False,
            disqualified=False,
        ).values('user_id').annotate(
            score_sum=Sum('score'),
            attempts=Count('id'),
        )
    }

    # Teg bo'yicha yig'amiz: ball yig'indisi va urinishlar (o'rtacha ball =
    # barcha urinishlar bo'yicha og'irlikli o'rtacha).
    tag_agg = {}
    for tag in tag_counts:
        tag_agg[tag] = {'score_sum': 0, 'attempts': 0}
    for uid, row in per_user.items():
        tag = tag_by_user.get(uid)
        if tag is None:
            continue
        tag_agg[tag]['score_sum'] += row['score_sum'] or 0
        tag_agg[tag]['attempts'] += row['attempts'] or 0

    rows = []
    for tag, count in tag_counts.items():
        agg = tag_agg.get(tag, {'score_sum': 0, 'attempts': 0})
        attempts = agg['attempts']
        avg_score = round(agg['score_sum'] / attempts, 1) if attempts else 0.0
        rows.append({
            'tag': tag,
            'student_count': count,
            'avg_score': avg_score,
            'total_attempts': attempts,
        })
    rows.sort(key=lambda r: (-r['avg_score'], -r['total_attempts']))
    return Response(rows)


# ─── T6. Yig'ma olimpiada eksporti ──────────────────────────────────────────


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@renderer_classes([JSONRenderer, _PassthroughXlsxRenderer, _PassthroughCsvRenderer])
def export_all_results(request, center_id):
    """GET /api/centers/<id>/export-all-results/?format=xlsx|csv

    Markazning barcha olimpiadalari bo'yicha bitta fayl. Har qator: o'quvchi
    ismi, olimpiada nomi, sana, ball, foiz. openpyxl bo'lsa .xlsx, aks holda
    CSV qaytaradi.
    """
    from attempts.models import TestAttempt

    center = get_object_or_404(EducationCenter, pk=center_id)
    if not user_can_manage_center(request.user, center):
        return Response({'detail': 'Forbidden'}, status=http_status.HTTP_403_FORBIDDEN)
    if not center.is_premium:
        return _premium_center_required()

    # T5: eksport amalini menejer logiga yozamiz.
    from .models import ManagerActivityLog
    from .services import log_manager_activity
    log_manager_activity(
        center, request.user, ManagerActivityLog.ACTION_EXPORT_DATA,
        description='Barcha natijalar eksporti',
    )

    fmt = (request.query_params.get('format') or 'xlsx').strip().lower()

    attempts = list(
        TestAttempt.objects
        .filter(olympiad__center=center, olympiad__is_deleted=False, disqualified=False)
        .select_related('user', 'olympiad')
        .order_by('olympiad__title', '-score', 'time_spent')
    )

    def _row(a):
        score = a.score or 0
        max_score = a.olympiad.max_score or 100
        pct = round((score / max_score) * 100, 1) if max_score else 0
        return [
            a.user.full_name or a.user.normalized_phone or '—',
            a.olympiad.title or '—',
            a.submitted_at.strftime('%Y-%m-%d %H:%M') if a.submitted_at else '',
            score,
            pct,
        ]

    headers = ["O'quvchi", 'Olimpiada', 'Sana', 'Ball', 'Foiz (%)']
    safe_name = ''.join(
        ch for ch in (center.name or 'markaz')
        if ch.isalnum() or ch in (' ', '_', '-')
    )[:50].strip().replace(' ', '_') or 'markaz'

    if fmt == 'csv':
        return _csv_response(headers, [_row(a) for a in attempts], f'olympy-natijalar-{safe_name}')

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return _csv_response(headers, [_row(a) for a in attempts], f'olympy-natijalar-{safe_name}')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Barcha natijalar'
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center')
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    for idx, a in enumerate(attempts, start=2):
        row = _row(a)
        ws.cell(row=idx, column=1, value=row[0])
        ws.cell(row=idx, column=2, value=row[1])
        ws.cell(row=idx, column=3, value=row[2]).alignment = center_align
        ws.cell(row=idx, column=4, value=row[3]).alignment = center_align
        ws.cell(row=idx, column=5, value=row[4]).alignment = center_align
    for i, w in enumerate([30, 32, 18, 10, 12], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="olympy-natijalar-{safe_name}.xlsx"'
    return response


def _csv_response(headers, rows, filename):
    import csv

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    # BOM bilan — Excel UTF-8'ni to'g'ri ochishi uchun.
    content = '﻿' + buf.getvalue()
    response = HttpResponse(content, content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    return response


# ─── T7. Markaz reyting dinamikasi ──────────────────────────────────────────


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def rating_history(request, center_id):
    """GET /api/centers/<id>/rating-history/?months=6

    Markazning oxirgi N oydagi reyting dinamikasi. Har oy uchun shu oydagi
    eng so'nggi yozuvning rank va score qiymati olinadi.
    Javob: [{month, rank, score}]
    """
    from .models import CenterRatingHistory

    center = get_object_or_404(EducationCenter, pk=center_id)
    if not user_can_manage_center(request.user, center):
        return Response({'detail': 'Forbidden'}, status=http_status.HTTP_403_FORBIDDEN)
    if not center.is_premium:
        return _premium_center_required()

    try:
        months = int(request.query_params.get('months') or 6)
    except (TypeError, ValueError):
        months = 6
    months = max(1, min(months, 24))

    now = timezone.now()
    cutoff = (now - timedelta(days=months * 31)).date()
    history = list(
        CenterRatingHistory.objects
        .filter(center=center, date__gte=cutoff)
        .order_by('date')
    )
    # Oy bo'yicha guruhlaymiz — har oyning eng so'nggi yozuvi.
    by_month = {}
    for h in history:
        key = h.date.strftime('%Y-%m')
        # order_by('date') — keyingisi (kattaroq sana) avvalgisini almashtiradi.
        by_month[key] = {
            'month': key,
            'rank': h.rank,
            'score': float(h.score or 0),
        }
    data = [by_month[k] for k in sorted(by_month.keys())]
    return Response(data)
