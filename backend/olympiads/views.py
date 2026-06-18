import csv

from django.db.models import Avg, Count, F, Max, Min
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from rest_framework import status as http_status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from accounts.models import AuditLog
from .models import Olympiad
from .serializers import OlympiadSerializer
from .services import (
    _queue_olympiad_summary,
    center_olympiad_limit_exceeded,
    event_readiness_errors,
    finalize_expired_active_olympiads,
    recompute_olympiad_ranks,
    user_can_manage_center_event,
    visible_events_filter,
)


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def olympiads_list_create(request):
    """GET /api/olympiads/  — visible olympiads/competitions.
    POST /api/olympiads/    — create draft event (manager/owner/admin).
    """
    if request.method == 'GET':
        # Celery worker yo'q muhitda (Render free tier) muddati o'tgan
        # olimpiadalar avtomatik FINISHED ga o'tmasdi va ACTIVE bo'lib
        # osilib qolardi — student ro'yxatda "Faol" deb ko'rib, ochmoqchi
        # bo'lganda "Olimpiada yakunlangan" deb rad etilardi, "Tugagan"
        # tabiga esa hech qachon o'tmasdi (tugagandek "yo'qolardi"). Ro'yxat
        # qaytarilishidan OLDIN muddati o'tgan ACTIVE'larni yopamiz, shunda
        # quyidagi queryset doim to'g'ri status (active/finished) qaytaradi.
        #
        # Throttle: CELERY_TASK_ALWAYS_EAGER (Redis yo'q, production free tier)
        # rejimida bu funksiya har GET so'rovda SINXRON ishlaydi va ro'yxat
        # yuklanishini sekinlashtiradi. Cache flag bilan 60 soniyada faqat
        # bir marta ishlatamiz — qolgan so'rovlar darhol javob qaytaradi.
        # Celery beat (Redis bor) muhitda ham bu kifoya, chunki beat baribir
        # davriy ishlaydi; bu yerda EAGER stsenariy uchun zaxira mexanizm.
        from django.conf import settings as dj_settings
        from django.core.cache import cache
        should_finalize = True
        if getattr(dj_settings, 'CELERY_TASK_ALWAYS_EAGER', False):
            # add() faqat kalit mavjud bo'lmaganda True qaytaradi (atomik) —
            # 60 soniya ichidagi keyingi so'rovlar False oladi va o'tkazib
            # yuboriladi.
            should_finalize = cache.add('finalize_expired_throttle', 1, 60)
        if should_finalize:
            try:
                finalize_expired_active_olympiads()
            except Exception:
                import logging
                logging.getLogger(__name__).warning(
                    'finalize_expired_active_olympiads skipped', exc_info=True,
                )
        # Avval `prefetch_related('questions')` to'liq Question modelidagi
        # barcha maydonlarni yuklab olardi (text, options, image, va h.k.).
        # List javobida faqat `question_ids` (id'lar massivi) va `max_score`
        # (questions__score yig'indisi) kerak. Shu sababli:
        #  1) `total_score` ni annotate orqali aggregate qilamiz — har bir
        #     olimpiada uchun questions'larni Python'da iteratsiya qilmasdan
        #     bitta SQL aggregate'da olinadi (max_score serializer'i shundan
        #     foydalanadi).
        #  2) Prefetch'ni `only('id')` bilan cheklab, faqat ID'larni olamiz —
        #     question_ids field uchun yetarli, lekin behuda kolonkalar
        #     yuklanmaydi.
        from django.db.models import OuterRef, Prefetch, Subquery, Sum, IntegerField
        from django.db.models.functions import Coalesce
        from questions.models import Question
        # Y10: avval `Sum('questions__score')` to'g'ridan-to'g'ri annotate
        # qilinardi va bu boshqa annotate'lar (Count('attempts'),
        # Avg('attempts__score')) bilan JOIN multipication hosil qilardi —
        # natija nohaq (questions × attempts marta) bo'lishi mumkin edi.
        # Endi Subquery orqali hisoblanadi va annotate'lar bir-biriga
        # ta'sir qilmaydi. Prefetch hali ham kerak — `question_ids`
        # serializer'i `obj.questions.all()` chaqiradi.
        total_score_sq = (
            Question.objects
            .filter(olympiads=OuterRef('pk'))
            .values('olympiads')
            .annotate(s=Sum('score'))
            .values('s')
        )
        queryset = (
            Olympiad.objects
            .prefetch_related(
                Prefetch('questions', queryset=Question.objects.only('id')),
            )
            .select_related('center')
            .annotate(
                participants_count=Count('attempts', distinct=True),
                avg_score_value=Avg('attempts__score'),
                total_score=Coalesce(
                    Subquery(total_score_sq, output_field=IntegerField()),
                    0,
                ),
            )
            .order_by('-created_at')
        )
        qs = queryset.filter(visible_events_filter(request.user), is_deleted=False).distinct()
        # Pagination: 500+ olimpiada bo'lishi mumkin, ayniqsa platform admin
        # uchun. Frontend grid'da hammasini bir martada ko'rsatish uchun
        # `?page_size=200` yuboradi — Default PageNumberPagination uni
        # e'tiborga olmasdi va 50 ta bilan chegaralanardi. LargePageNumberPagination
        # `page_size_query_param='page_size'` va `max_page_size=200` bilan
        # frontend so'rovini hurmat qiladi.
        from olympy_api.pagination import LargePageNumberPagination
        paginator = LargePageNumberPagination()
        page = paginator.paginate_queryset(qs, request)
        if page is not None:
            return paginator.get_paginated_response(OlympiadSerializer(page, many=True).data)
        return Response(OlympiadSerializer(qs, many=True).data)

    serializer = OlympiadSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    center = serializer.validated_data['center']
    questions = serializer.validated_data.pop('questions', None)
    if not user_can_manage_center_event(request.user, center):
        return Response({'detail': "Sizda bu markaz uchun tadbir yaratish huquqi yo'q"},
                        status=http_status.HTTP_403_FORBIDDEN)
    # Bepul markaz limiti: joriy oyda max FREE_OLYMPIAD_MONTHLY_LIMIT ta.
    # Platform admin uchun limit qo'llanilmaydi.
    if not request.user.is_platform_admin and center_olympiad_limit_exceeded(center):
        from django.conf import settings
        limit = getattr(settings, 'FREE_OLYMPIAD_MONTHLY_LIMIT', 2)
        return Response(
            {
                'detail': (
                    f'Bepul rejimda oyiga {limit} ta olimpiada yaratish mumkin. '
                    'Cheksiz yaratish uchun premium oling.'
                ),
                'upgrade_required': True,
            },
            status=http_status.HTTP_403_FORBIDDEN,
        )
    # O4: olimpiada yaratish va savollar biriktirish bitta transaction'da —
    # `serializer.save()` muvaffaqiyatli bo'lib, `questions.set()` xato bersa
    # (masalan, savol ID xato), olimpiada savollarsiz qolib ketardi.
    from django.db import transaction
    with transaction.atomic():
        olympiad = serializer.save(
            created_by=request.user,
            status=Olympiad.STATUS_DRAFT,
        )
        if questions is not None:
            olympiad.questions.set(questions)
    AuditLog.log(request, 'olympiad_create', target=olympiad, extra={
        'title': olympiad.title,
        'center_id': olympiad.center_id,
        'event_type': olympiad.event_type,
    })
    return Response(OlympiadSerializer(olympiad).data,
                    status=http_status.HTTP_201_CREATED)


@api_view(['PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def olympiad_detail(request, olympiad_id):
    """PATCH /api/olympiads/{id}/ — update draft/inactive event fields/questions.
    DELETE /api/olympiads/{id}/ — delete draft/inactive event.
    """
    olympiad = get_object_or_404(Olympiad, pk=olympiad_id)
    if not user_can_manage_center_event(request.user, olympiad.center):
        return Response({'detail': 'Forbidden'},
                        status=http_status.HTTP_403_FORBIDDEN)
    if request.method == 'DELETE':
        if olympiad.status == Olympiad.STATUS_ACTIVE:
            return Response(
                {'detail': "Faol tadbirni o'chirish mumkin emas. Avval uni nofaol qiling."},
                status=http_status.HTTP_400_BAD_REQUEST,
            )
        olympiad.is_deleted = True
        olympiad.save(update_fields=['is_deleted'])
        AuditLog.log(request, 'olympiad_delete', target=olympiad, extra={
            'title': olympiad.title,
            'center_id': olympiad.center_id,
        })
        return Response({'detail': "Tadbir muvaffaqiyatli o'chirildi"}, status=http_status.HTTP_200_OK)

    if olympiad.status not in [Olympiad.STATUS_DRAFT, Olympiad.STATUS_INACTIVE]:
        return Response(
            {'detail': "Faqat draft yoki nofaol tadbirni tahrirlash mumkin"},
            status=http_status.HTTP_400_BAD_REQUEST,
        )

    serializer = OlympiadSerializer(olympiad, data=request.data, partial=True)
    serializer.is_valid(raise_exception=True)
    questions = serializer.validated_data.pop('questions', None)
    # O4: PATCH ham atomic — savollar tugamasdan qisman saqlanish bo'lmasin.
    from django.db import transaction
    with transaction.atomic():
        olympiad = serializer.save()
        if questions is not None:
            olympiad.questions.set(questions)
    return Response(OlympiadSerializer(olympiad).data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def publish_olympiad(request, olympiad_id):
    """POST /api/olympiads/{id}/publish/ — flip status to active and notify."""
    olympiad = get_object_or_404(Olympiad, pk=olympiad_id)
    if not user_can_manage_center_event(request.user, olympiad.center):
        return Response({'detail': 'Forbidden'},
                        status=http_status.HTTP_403_FORBIDDEN)
    if olympiad.status not in [Olympiad.STATUS_DRAFT, Olympiad.STATUS_INACTIVE]:
        return Response(
            {'detail': 'Faqat draft yoki nofaol tadbirni faollashtirish mumkin'},
            status=http_status.HTTP_400_BAD_REQUEST,
        )
    readiness_errors = event_readiness_errors(olympiad)
    if readiness_errors:
        return Response(
            {'detail': 'Tadbir hali tayyor emas', 'errors': readiness_errors},
            status=http_status.HTTP_400_BAD_REQUEST,
        )
    # Avval xabar har gal publish'da yuborilardi — INACTIVE → ACTIVE qayta
    # nashr qilinganda ham. Endi faqat DRAFT'dan birinchi marta nashr
    # qilinayotganda xabar yuboriladi: studentlar ikki marta xabar olmaydi.
    is_first_publish = olympiad.status == Olympiad.STATUS_DRAFT
    olympiad.status = Olympiad.STATUS_ACTIVE
    olympiad.save(update_fields=['status'])

    if is_first_publish and olympiad.event_type in (
        Olympiad.EVENT_TYPE_COMPETITION,
        Olympiad.EVENT_TYPE_OLYMPIAD,
    ):
        # Lazy import: avoid circular dependency.
        from centers.models import CenterMembership
        from notifications.services import send_olympiad_published_bulk
        # Olimpiada (public) bo'lsa-da, push spam'ni oldini olish uchun
        # xabar faqat shu markazning approved studentlariga yuboriladi —
        # boshqa markaz a'zolari va platforma userlari uchun olimpiada
        # baribir feed/listda paydo bo'ladi.
        approved_students = CenterMembership.objects.filter(
            center=olympiad.center,
            role=CenterMembership.ROLE_STUDENT,
            status=CenterMembership.STATUS_APPROVED,
        ).select_related('user')
        try:
            send_olympiad_published_bulk(
                [m.user for m in approved_students],
                olympiad,
                olympiad.center,
            )
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning('Notification send failed: %s', e)
    return Response(OlympiadSerializer(olympiad).data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def finish_olympiad(request, olympiad_id):
    """POST /api/olympiads/{id}/finish/ — flip status to finished."""
    olympiad = get_object_or_404(Olympiad, pk=olympiad_id)
    if not user_can_manage_center_event(request.user, olympiad.center):
        return Response({'detail': 'Forbidden'},
                        status=http_status.HTTP_403_FORBIDDEN)
    if olympiad.status != Olympiad.STATUS_ACTIVE:
        return Response({'detail': "Faqat faol tadbirni yakunlash mumkin"},
                        status=http_status.HTTP_400_BAD_REQUEST)
    from django.db import transaction
    with transaction.atomic():
        olympiad.status = Olympiad.STATUS_FINISHED
        olympiad.save(update_fields=['status'])
        # Rank'larni manualda yakunlashda ham qayta hisoblaymiz — submit
        # paytida yangilanmaydi (DB yukini kamaytirish), shu sababli
        # yakunlash paytida bir martalik bulk update kifoya.
        recompute_olympiad_ranks(olympiad)
        # Markazga bog'liq olimpiada bo'lsa — menejer/ustozlarga xulosa
        # yuboramiz (commit'dan keyin, asinxron). Status tekshiruvi (yuqorida
        # `STATUS_ACTIVE` shart) tufayli xabar bir marta yuboriladi.
        if olympiad.center_id:
            _queue_olympiad_summary(olympiad.pk)
    return Response(OlympiadSerializer(olympiad).data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def deactivate_olympiad(request, olympiad_id):
    """POST /api/olympiads/{id}/deactivate/ — pause an active event for editing."""
    olympiad = get_object_or_404(Olympiad, pk=olympiad_id)
    if not user_can_manage_center_event(request.user, olympiad.center):
        return Response({'detail': 'Forbidden'},
                        status=http_status.HTTP_403_FORBIDDEN)
    if olympiad.status != Olympiad.STATUS_ACTIVE:
        return Response({'detail': 'Faqat faol tadbirni nofaollashtirish mumkin'},
                        status=http_status.HTTP_400_BAD_REQUEST)
    from django.db import transaction
    from attempts.models import TestAttempt, TestSession
    from attempts.session_utils import score_session_answers
    # Avval olympiad inaktiv qilinganda hali test yechayotgan studentlarning
    # javoblari yo'qolar va bo'sh attempt bilan to'ldirilardi. Endi admin
    # oldindan ogohlantiriladi: hech bo'lmaganda bitta faol session bo'lsa
    # va `force=true` yuborilmagan bo'lsa, 400 qaytariladi. Force bilan
    # yuborilsa session'dagi mavjud javoblar bo'yicha ball hisoblanadi
    # (bo'sh attempt o'rniga) — student haqiqiy javoblarini yo'qotmaydi.
    force = str(request.data.get('force') or '').lower() in ('1', 'true', 'yes')
    has_active_sessions = TestSession.objects.filter(
        olympiad=olympiad,
        status=TestSession.STATUS_ACTIVE,
    ).exists()
    if has_active_sessions and not force:
        return Response(
            {
                'detail': (
                    "Faol ishtirokchilar bor — olimpiadani to'xtatib bo'lmaydi. "
                    "Baribir to'xtatish uchun {\"force\": true} bilan qayta yuboring."
                ),
                'active_sessions': True,
            },
            status=http_status.HTTP_400_BAD_REQUEST,
        )
    # Butun deaktivatsiya bloki atomic — yarmida xatolik bo'lsa session
    # COMPLETED bo'lib qoladi, lekin attempt yaratilmaydi degan
    # nomuvofiqlikni oldini olamiz.
    with transaction.atomic():
        active_sessions = list(TestSession.objects.filter(
            olympiad=olympiad,
            status=TestSession.STATUS_ACTIVE,
        ).select_related('user'))
        TestSession.objects.filter(
            olympiad=olympiad,
            status=TestSession.STATUS_ACTIVE,
        ).update(status=TestSession.STATUS_COMPLETED)
        # Faol sessiyalarning egalari uchun mavjud bo'lmagan attempt'larni
        # session'dagi javoblar bo'yicha hisoblangan natija bilan yaratamiz.
        # Sessionda javoblar saqlanmaydi (faqat question_order/option_orders),
        # shu sababli answers={} bilan keladi — bu holatda blank qoladi.
        # Lekin agar kelajakda sessionga answers qo'shilsa, shu kod
        # avtomatik foydalanadi.
        if active_sessions:
            existing_user_ids = set(TestAttempt.objects.filter(
                olympiad=olympiad,
                user_id__in=[s.user_id for s in active_sessions],
            ).values_list('user_id', flat=True))
            to_create = []
            for s in active_sessions:
                if s.user_id in existing_user_ids:
                    continue
                # Sessionda saqlangan javoblar yo'q — frontend localStorage'da
                # saqlaydi va submit'da yuboradi. Force deactivate'da bu javoblar
                # backend'ga yetib bormaydi, shu sababli boshlangan vaqtdan
                # hozirgacha bo'lgan time_spent ham yozamiz — student keyin
                # statistikada "qatnashgan" deb ko'rinadi.
                scored = score_session_answers(s, olympiad, {})
                time_spent = max(0, int(
                    (timezone.now() - s.started_at).total_seconds()
                )) if s.started_at else 0
                to_create.append(TestAttempt(
                    user=s.user,
                    olympiad=olympiad,
                    answers={},
                    score=scored.get('score', 0),
                    correct_count=scored.get('correct', 0),
                    wrong_count=scored.get('wrong', 0),
                    total_questions=scored.get('total', 0),
                    time_spent=time_spent,
                    rank=None,
                ))
            if to_create:
                TestAttempt.objects.bulk_create(to_create)
        olympiad.status = Olympiad.STATUS_INACTIVE
        olympiad.save(update_fields=['status'])
    return Response(OlympiadSerializer(olympiad).data)


def _safe_olympiad_filename(olympiad):
    """Fayl nomi uchun xavfsiz title ("Matematika 2024" -> "Matematika_2024")."""
    safe_title = ''.join(
        ch for ch in (olympiad.title or 'olimpiada')
        if ch.isalnum() or ch in (' ', '_', '-')
    )[:60].strip() or 'olimpiada'
    return safe_title.replace(' ', '_')


def _export_attempts_queryset(olympiad):
    """Eksport uchun saralangan attempt'lar (faqat topshirilgan, rank bo'yicha).

    `submitted_at` modelda `auto_now_add` — har doim to'ldirilgan, shu sababli
    "topshirilgan" sharti diskvalifikatsiya bo'lmagan attempt'lar bilan teng.
    """
    from attempts.models import TestAttempt
    return (
        TestAttempt.objects
        .filter(olympiad=olympiad, disqualified=False)
        .select_related('user')
        .order_by('rank', '-score', 'time_spent')
    )


def _export_row_values(idx, attempt):
    """Bitta attempt uchun ustun qiymatlari (CSV/XLSX/PDF — bir xil tartib).

    `idx` — 1 dan boshlanadigan tartib raqami (rank bo'sh bo'lsa fallback).
    """
    user = attempt.user
    full_name = getattr(user, 'full_name', '') or '—'
    username = getattr(user, 'username', '') or '—'
    total = attempt.total_questions or 0
    answered = (attempt.correct_count or 0) + (attempt.wrong_count or 0)
    pct = round((attempt.correct_count / total) * 100) if total else 0
    minutes = round((attempt.time_spent or 0) / 60.0, 1)
    return [
        attempt.rank or idx,
        full_name,
        username,
        attempt.score,
        attempt.correct_count,
        attempt.wrong_count,
        total,
        f'{pct}%',
        f'{minutes} daq',
    ]


EXPORT_HEADERS = [
    '#', 'Ism', 'Username', 'Ball', "To'g'ri", "Noto'g'ri", 'Jami', '%', 'Vaqt',
]


def _export_csv(olympiad, attempts):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = (
        f'attachment; filename="olympy-{_safe_olympiad_filename(olympiad)}-{olympiad.id}-results.csv"'
    )
    # UTF-8 BOM — Excel CSV ni avtomatik UTF-8 sifatida tan oladi.
    response.write('﻿')
    writer = csv.writer(response)
    writer.writerow(EXPORT_HEADERS)
    for idx, a in enumerate(attempts, start=1):
        writer.writerow(_export_row_values(idx, a))
    return response


def _export_xlsx(olympiad, attempts):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return Response(
            {'detail': "Excel eksport moduli o'rnatilmagan"},
            status=http_status.HTTP_503_SERVICE_UNAVAILABLE,
        )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Natijalar'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center')
    for col_idx, header in enumerate(EXPORT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for idx, a in enumerate(attempts, start=1):
        values = _export_row_values(idx, a)
        row_idx = idx + 1
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx != 2:  # Ism ustunidan tashqari hammasi markazga
                cell.alignment = center_align

    # Eng pastda jami qatnashchilar soni.
    total_row = len(attempts) + 2
    label_cell = ws.cell(row=total_row, column=1, value='Jami qatnashchilar:')
    label_cell.font = Font(bold=True)
    count_cell = ws.cell(row=total_row, column=4, value=len(attempts))
    count_cell.font = Font(bold=True)
    count_cell.alignment = center_align

    column_widths = [6, 28, 20, 8, 9, 10, 8, 8, 10]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width
    ws.freeze_panes = 'A2'

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = (
        f'attachment; filename="olympy-{_safe_olympiad_filename(olympiad)}-{olympiad.id}-results.xlsx"'
    )
    return response


def _export_pdf(olympiad, attempts):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
        )
    except ImportError:
        return Response(
            {'detail': "PDF eksport moduli o'rnatilmagan"},
            status=http_status.HTTP_503_SERVICE_UNAVAILABLE,
        )

    from io import BytesIO
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        leftMargin=1.2 * cm, rightMargin=1.2 * cm,
        title=f'{olympiad.title} — natijalar',
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CenterTitle', parent=styles['Title'], fontSize=20, spaceAfter=4,
        textColor=colors.HexColor('#1e1b4b'),
    )
    sub_style = ParagraphStyle(
        'SubTitle', parent=styles['Heading2'], fontSize=13, spaceAfter=2,
        textColor=colors.HexColor('#4f46e5'),
    )
    meta_style = ParagraphStyle(
        'Meta', parent=styles['Normal'], fontSize=10,
        textColor=colors.HexColor('#555555'),
    )

    center_name = (olympiad.center.name if olympiad.center else 'Olympy') or 'Olympy'
    now = timezone.localtime(timezone.now())
    elements = [
        Paragraph(center_name, title_style),
        Paragraph(olympiad.title or 'Olimpiada', sub_style),
        Paragraph(f"Sana: {now.strftime('%Y-%m-%d %H:%M')}", meta_style),
        Spacer(1, 0.5 * cm),
    ]

    data = [EXPORT_HEADERS]
    for idx, a in enumerate(attempts, start=1):
        data.append([str(v) for v in _export_row_values(idx, a)])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4f46e5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 1), (2, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cccccc')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.6 * cm))
    elements.append(Paragraph(
        f"Jami qatnashchilar: {len(attempts)}", meta_style,
    ))
    elements.append(Spacer(1, 0.3 * cm))
    elements.append(Paragraph(
        f"Yuklab olindi: {now.strftime('%Y-%m-%d %H:%M')} · olympy.uz",
        ParagraphStyle('Footer', parent=meta_style, fontSize=8,
                       textColor=colors.HexColor('#888888')),
    ))

    doc.build(elements)
    buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="olympy-{_safe_olympiad_filename(olympiad)}-{olympiad.id}-results.pdf"'
    )
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_olympiad_results(request, olympiad_id):
    """GET /api/olympiads/{id}/export/?format=csv|xlsx|pdf — natijalar fayli.

    - csv (default): yengil, har joyda ochiladigan jadval (orqaga moslik —
      eski "Natijalarni yuklab olish" tugmasi format'siz so'rab CSV oladi).
    - xlsx: formatlangan Excel jadvali (header rang, ustun kengligi, jami).
    - pdf: rasmiy natijalar varaqasi (markaz nomi, olimpiada, sana, jadval).

    Ruxsat: center owner/manager/teacher va platform admin. Bundan tashqari
    XLSX/PDF eksport faqat Plus/Pro obuna (yoki lifetime premium) markazlar
    uchun — bepul/Standart markaz 403 oladi (CSV barcha uchun ochiq qoladi,
    chunki u allaqachon mavjud bepul imkoniyat edi).
    """
    olympiad = get_object_or_404(
        Olympiad.objects.select_related('center'),
        pk=olympiad_id,
    )
    if not user_can_manage_center_event(request.user, olympiad.center):
        return Response({'detail': 'Forbidden'},
                        status=http_status.HTTP_403_FORBIDDEN)

    fmt = (request.query_params.get('format') or 'csv').lower()
    if fmt not in ('csv', 'xlsx', 'pdf'):
        fmt = 'csv'

    # Plus/Pro tekshiruvi faqat formatlangan eksport (xlsx/pdf) uchun. Platform
    # admin uchun limit qo'llanilmaydi.
    if fmt in ('xlsx', 'pdf') and not request.user.is_platform_admin:
        from billing.services import SubscriptionService
        if olympiad.center and not SubscriptionService(olympiad.center).can_export_results():
            return Response(
                {
                    'detail': 'Natijalarni yuklab olish Plus yoki Pro obunasi talab qiladi.',
                    'upgrade_required': True,
                },
                status=http_status.HTTP_403_FORBIDDEN,
            )

    attempts = list(_export_attempts_queryset(olympiad))

    if fmt == 'xlsx':
        return _export_xlsx(olympiad, attempts)
    if fmt == 'pdf':
        return _export_pdf(olympiad, attempts)
    return _export_csv(olympiad, attempts)


# Orqaga moslik: eski import nomi (urls.py va boshqa modullar `export_results`
# nomiga tayanishi mumkin). Yangi funksiya format'siz so'rovda CSV qaytaradi —
# eski xulq-atvor saqlanadi.
export_results = export_olympiad_results


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def code_submissions(request, olympiad_id):
    """GET /api/olympiads/{id}/code-submissions/ — IT (kod) javoblari ro'yxati.

    Faqat center owner/manager/teacher va platform admin uchun. Olimpiadaga
    yuborilgan barcha kod javoblar + AI tavsiyalari (ManagerDashboard "Kod
    javoblari" tabi shu endpoint'ga ulanadi).
    """
    olympiad = get_object_or_404(
        Olympiad.objects.select_related('center'),
        pk=olympiad_id,
    )
    if not user_can_manage_center_event(request.user, olympiad.center):
        return Response({'detail': 'Forbidden'},
                        status=http_status.HTTP_403_FORBIDDEN)

    from attempts.models import CodeSubmission
    from attempts.serializers import CodeSubmissionSerializer

    submissions = (
        CodeSubmission.objects
        .filter(attempt__olympiad=olympiad)
        .select_related('attempt', 'attempt__user', 'question')
        .order_by('question_id', 'attempt__user_id')
    )
    return Response(CodeSubmissionSerializer(submissions, many=True).data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def olympiad_stats(request, olympiad_id):
    """GET /api/olympiads/{id}/stats/ — agregat statistika.

    Owner/Manager dashboard'da ishlatiladi: ishtirokchilar soni, o'rtacha
    ball, eng yuqori/eng past ball, to'liq yechganlar foizi, o'rtacha
    sarflangan vaqt.
    """
    olympiad = get_object_or_404(
        Olympiad.objects.select_related('center'),
        pk=olympiad_id,
    )
    if not user_can_manage_center_event(request.user, olympiad.center):
        return Response({'detail': 'Forbidden'},
                        status=http_status.HTTP_403_FORBIDDEN)

    from attempts.models import TestAttempt
    aggregates = TestAttempt.objects.filter(olympiad=olympiad).aggregate(
        participants=Count('id'),
        avg_score=Avg('score'),
        max_score=Max('score'),
        min_score=Min('score'),
        avg_time=Avg('time_spent'),
        avg_correct=Avg('correct_count'),
    )
    participants = aggregates.get('participants') or 0
    # To'liq yechganlar = score == max_score yoki score >= 90 emas; aniq
    # ta'rif: barcha savollarga javob bergan attempts soni.
    full_complete = (
        TestAttempt.objects.filter(
            olympiad=olympiad,
            total_questions__gt=0,
        )
        .annotate(answered=F('correct_count') + F('wrong_count'))
        .filter(answered__gte=F('total_questions'))
        .count()
        if participants else 0
    )
    full_complete_pct = (
        round((full_complete / participants) * 100, 1) if participants else 0.0
    )
    return Response({
        'olympiad_id': olympiad.id,
        'title': olympiad.title,
        'participants': participants,
        'average_score': round(aggregates.get('avg_score') or 0, 1),
        'max_score': aggregates.get('max_score') or 0,
        'min_score': aggregates.get('min_score') or 0,
        'average_time_seconds': round(aggregates.get('avg_time') or 0, 1),
        'average_correct': round(aggregates.get('avg_correct') or 0, 1),
        'full_complete_count': full_complete,
        'full_complete_percent': full_complete_pct,
    })
