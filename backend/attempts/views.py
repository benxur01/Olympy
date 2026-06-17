from datetime import timedelta

from django.db import IntegrityError, transaction
from django.db.models import Avg, Count, Max, Min, Q
from django.db.models.functions import TruncMonth
from django.shortcuts import get_object_or_404
from django.utils import timezone
from rest_framework import status as http_status
from rest_framework.decorators import api_view, permission_classes, throttle_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.throttling import ScopedRateThrottle
from rest_framework.views import APIView

from centers.models import CenterMembership, EducationCenter
from olympiads.models import Olympiad
from olympiads.services import (
    maybe_finish_expired_olympiad,
    user_can_manage_center_event,
    user_can_participate_in_event,
)


# O10: `_recompute_center_rating` olib tashlandi — submit_attempt'da
# chaqirilmasdi (DB yukini kamaytirish uchun) va hech qaerda ham
# foydalanilmasdi. Markaz reytingi keyinroq cron orqali tiklash kerak
# bo'lsa, alohida `services.py` faylida qayta yoziladi.

from django.http import HttpResponse

from questions.grading import RESULT_CORRECT, RESULT_WRONG, grade_answer

from .certificates import render_certificate_png
from .models import TestAttempt, TestSession
from .serializers import SubmitAttemptSerializer, TestAttemptSerializer
from .session_utils import (
    SUBMIT_GRACE_SECONDS,
    score_session_answers,
    session_is_expired,
)


def _extract_review_chosen(chosen, q_type):
    """Saqlangan javob payload'idan review uchun xom qiymatni ajratadi.

    score_session_answers/_extract_chosen bilan bir xil shartnoma — yangi
    turlar obyekt-shaklli ({"chosen_idx"..}, {"selected"..}, {"text"..})
    bo'lishi mumkin; eski skalar formatlar ham qo'llab-quvvatlanadi.
    """
    if isinstance(chosen, dict):
        if q_type in ('mcq', 'yes_no'):
            return chosen.get('chosen_idx')
        if q_type == 'multiple_select':
            return chosen.get('selected')
        if q_type in ('fill_blank', 'essay'):
            return chosen.get('text')
        if q_type == 'fill_blanks':
            if 'blanks' in chosen:
                return chosen.get('blanks')
            return chosen
    return chosen


def _build_attempt_mistakes(attempt, olympiad, answers):
    """O4: Attempt bo'yicha xato savollar ro'yxatini tuzadi (AI tahlil uchun).

    Avval faqat int (mcq indeks) javoblar tushunilardi — multiple_select,
    fill_blank kabi turlar jimgina skip bo'lardi. Endi barcha turlar
    questions.grading.grade_answer orqali izchil baholanadi; variant indeksli
    turlarda sessiyadagi option_orders bilan de-shuffle qilinadi.
    """
    from questions.grading import RESULT_WRONG, grade_answer

    from .session_utils import _deshuffle_index, _deshuffle_multi

    mistakes = []
    answers = answers or {}
    option_orders = {}
    session = TestSession.objects.filter(
        user=attempt.user, olympiad=olympiad,
    ).only('option_orders').first()
    if session:
        option_orders = session.option_orders or {}

    for q in olympiad.questions.all().order_by('id'):
        chosen = answers.get(str(q.id))
        if chosen is None:
            chosen = answers.get(q.id)
        if chosen is None:
            continue
        q_type = getattr(q, 'question_type', 'mcq') or 'mcq'
        # Kod va essay avtomatik baholanmaydi — mistakes ro'yxatiga kirmaydi.
        if q_type in ('code', 'essay'):
            continue
        chosen = _extract_review_chosen(chosen, q_type)
        options = list(q.options or [])
        order = option_orders.get(str(q.id)) or list(range(len(options)))
        if q_type in ('mcq', 'yes_no'):
            chosen = _deshuffle_index(chosen, order)
        elif q_type == 'multiple_select':
            chosen = _deshuffle_multi(chosen, order)
        if grade_answer(q, chosen) != RESULT_WRONG:
            continue
        mistakes.append({
            'text': (q.text or '')[:200],
            'correct_answer': (
                q.correct_answer if q_type in ('mcq', 'yes_no')
                else getattr(q, 'correct_text', '') or q.correct_answer
            ),
            'chosen_answer': chosen,
        })
        if len(mistakes) >= 6:
            break
    return mistakes


def _trigger_attempt_ai_analysis(attempt, olympiad, answers):
    """O4: AttemptAIAnalysis pending yozuv yaratib, AI call'ni Celery'da bajaradi.

    Submit latency'ni bloklamaslik uchun butun AI generatsiya asinxron
    vazifada (Celery) ishlaydi.
    """
    from .models import AttemptAIAnalysis
    from .tasks import generate_attempt_ai_analysis_task

    # Allaqachon mavjud bo'lsa qayta yaratmaymiz (idempotent).
    obj, created = AttemptAIAnalysis.objects.get_or_create(
        attempt=attempt,
        defaults={'status': AttemptAIAnalysis.STATUS_PENDING},
    )
    if not created and obj.status == AttemptAIAnalysis.STATUS_READY:
        return

    generate_attempt_ai_analysis_task.delay(attempt.id)


def _save_code_submissions(attempt, olympiad, code_answers):
    """IT (kod) savollariga yuborilgan javoblarni CodeSubmission'ga saqlaydi
    va AI baholashni Celery asinxron vazifasida ishga tushiradi.

    `code_answers` — { "<question_id>": {"code": "...", "language": "..."} }.
    Faqat olimpiadaga biriktirilgan, question_type='code' savollar saqlanadi.
    """
    if not code_answers:
        return

    import uuid

    from django.core.cache import cache

    from questions.judge0_service import is_supported
    from questions.models import Question
    from questions.tasks import run_code_async_task
    from .models import CodeSubmission
    from .tasks import review_code_submissions_task

    # Olimpiadaning kod savollarini bir so'rovda olamiz.
    code_questions = {
        q.id: q
        for q in olympiad.questions.filter(
            question_type=Question.QUESTION_TYPE_CODE,
        )
    }
    if not code_questions:
        return

    to_review = []
    for raw_qid, payload in (code_answers or {}).items():
        try:
            qid = int(raw_qid)
        except (TypeError, ValueError):
            continue
        question = code_questions.get(qid)
        if not question:
            continue
        code = str((payload or {}).get('code') or '')
        language = str((payload or {}).get('language') or '').strip().lower()
        if not language:
            language = question.programming_language or ''
        submission, _ = CodeSubmission.objects.update_or_create(
            attempt=attempt,
            question=question,
            defaults={
                'submitted_code': code,
                'code_language': language,
                # Qayta submit (update) holatida eski natija qolib ketmasin —
                # Judge0 qayta tekshirib yangilaydi.
                'all_tests_passed': None,
            },
        )
        if code.strip():
            to_review.append(submission.id)
            # Avtomatik ball uchun: kodni Judge0 test caslari bo'yicha
            # tekshirib, natijani (all_tests_passed) shu submission yozuviga
            # yozamiz. Til qo'llab-quvvatlanmasa Judge0 baribir xato qaytaradi,
            # shu sababli faqat supported tillarni yuboramiz (aks holda
            # all_tests_passed None qoladi — ball berilmaydi).
            if is_supported(language):
                task_id = str(uuid.uuid4())
                cache.set(f"run_code:task:{task_id}", {'status': 'PENDING'}, timeout=300)
                run_code_async_task.delay(
                    task_id, code, language, '', question.id,
                    submission_id=submission.id,
                )

    if to_review:
        review_code_submissions_task.delay(to_review)


def _user_can_manage_olympiad(user, olympiad):
    """Olympiad uchun manager/owner/teacher/admin huquqi tekshiruvi."""
    if user.is_platform_admin:
        return True
    center = olympiad.center
    if not center:
        return False
    if center.owner_id == user.id:
        return True
    return CenterMembership.objects.filter(
        user=user,
        center=center,
        role__in=[
            CenterMembership.ROLE_MANAGER,
            CenterMembership.ROLE_OWNER,
            CenterMembership.ROLE_TEACHER,
        ],
        status=CenterMembership.STATUS_APPROVED,
    ).exists()


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_olympiad_results_xlsx(request, olympiad_id):
    """GET /api/manager/olympiads/<id>/export/ — natijalarni .xlsx faylga eksport.

    Ustunlar: O'rin, O'quvchi ismi, Telefon, Ball (%), To'g'ri, Noto'g'ri,
    Vaqt (daqiqa), Sana. Faqat shu markaz manager/owner/teacher/admin uchun.
    """
    olympiad = get_object_or_404(
        Olympiad.objects.select_related('center'),
        pk=olympiad_id,
    )
    if not _user_can_manage_olympiad(request.user, olympiad):
        return Response({'detail': 'Forbidden'}, status=http_status.HTTP_403_FORBIDDEN)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return Response(
            {'detail': "openpyxl o'rnatilmagan"},
            status=http_status.HTTP_503_SERVICE_UNAVAILABLE,
        )

    attempts = list(
        TestAttempt.objects
        .filter(olympiad=olympiad, disqualified=False)
        .select_related('user')
        .order_by('-score', 'time_spent', 'submitted_at')
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Natijalar"

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center')
    headers = [
        "O'rin",
        "O'quvchi ismi",
        'Telefon',
        'Ball (%)',
        "To'g'ri",
        "Noto'g'ri",
        'Vaqt (daqiqa)',
        'Sana',
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for idx, attempt in enumerate(attempts, start=1):
        user = attempt.user
        full_name = getattr(user, 'full_name', '') or '—'
        phone = getattr(user, 'normalized_phone', '') or getattr(user, 'phone', '') or '—'
        time_minutes = round((attempt.time_spent or 0) / 60.0, 1)
        submitted_date = (
            attempt.submitted_at.strftime('%Y-%m-%d %H:%M')
            if attempt.submitted_at else ''
        )
        row_idx = idx + 1
        ws.cell(row=row_idx, column=1, value=idx).alignment = center_align
        ws.cell(row=row_idx, column=2, value=full_name)
        ws.cell(row=row_idx, column=3, value=phone)
        ws.cell(row=row_idx, column=4, value=attempt.score).alignment = center_align
        ws.cell(row=row_idx, column=5, value=attempt.correct_count).alignment = center_align
        ws.cell(row=row_idx, column=6, value=attempt.wrong_count).alignment = center_align
        ws.cell(row=row_idx, column=7, value=time_minutes).alignment = center_align
        ws.cell(row=row_idx, column=8, value=submitted_date).alignment = center_align

    column_widths = [8, 30, 18, 12, 10, 12, 14, 18]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width
    ws.freeze_panes = 'A2'

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    safe_title = ''.join(
        ch for ch in (olympiad.title or 'olimpiada')
        if ch.isalnum() or ch in (' ', '_', '-')
    )[:60].strip() or 'olimpiada'
    safe_title = safe_title.replace(' ', '_')
    response['Content-Disposition'] = (
        f'attachment; filename="olympy-{safe_title}-{olympiad.id}-results.xlsx"'
    )
    return response


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@throttle_classes([ScopedRateThrottle])
def submit_attempt(request):
    """POST /api/attempts/ — student submits answers, server scores them.

    Enforces event access rules: public olympiads accept any authenticated
    participant, center competitions require an approved student membership
    at the event center. The event must be active, and one user can only
    submit once per event.
    """
    serializer = SubmitAttemptSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)

    # Celery worker Render free tier'da ishlamasligi mumkin — lazy ravishda
    # muddati o'tgan olimpiadalarni yopib qo'yamiz. Bu atomic transaction
    # dan TASHQARIDA bo'lishi kerak, aks holda ichkarida olingan olimpiad
    # qatori bilan to'qnashuv yuzaga keladi.
    try:
        _peek_olympiad = Olympiad.objects.filter(
            pk=serializer.validated_data['olympiad'],
        ).first()
        maybe_finish_expired_olympiad(_peek_olympiad)
    except Exception:
        import logging
        logging.getLogger(__name__).exception(
            'maybe_finish_expired_olympiad failed for olympiad=%s',
            serializer.validated_data.get('olympiad'),
        )

    with transaction.atomic():
        # Olimpiad qatorini lock qilmaymiz — bu butun olimpiada bo'yicha
        # submit'larni navbatga qo'yardi (100+ student bir vaqtda submit
        # qilsa katta latency). Yagona-attempt cheklovi va sessiya
        # tutqichi (TestSession.select_for_update) yetarli.
        olympiad = get_object_or_404(
            Olympiad.objects,
            pk=serializer.validated_data['olympiad'],
        )

        # Soft-delete qilingan olimpiadaga submit qabul qilinmaydi —
        # frontend'da ko'rinmasligi kerak, ammo URL'ni bilgan abuser
        # to'g'ridan-to'g'ri API'ga POST yuborishi mumkin. 404 — chunki
        # foydalanuvchi nuqtai nazaridan olimpiada mavjud emas.
        if olympiad.is_deleted:
            return Response(
                {'detail': 'Olimpiada topilmadi.'},
                status=http_status.HTTP_404_NOT_FOUND,
            )

        if TestAttempt.objects.filter(user=request.user, olympiad=olympiad).exists():
            return Response(
                {'detail': "Siz bu olimpiadaga allaqachon qatnashgansiz"},
                status=http_status.HTTP_400_BAD_REQUEST,
            )

        if olympiad.status != Olympiad.STATUS_ACTIVE:
            return Response({'detail': "Olimpiada faol emas"},
                            status=http_status.HTTP_400_BAD_REQUEST)
        now = timezone.now()
        if olympiad.start_datetime and now < olympiad.start_datetime:
            return Response({'detail': 'Olimpiada hali boshlanmagan'},
                            status=http_status.HTTP_400_BAD_REQUEST)
        end_time = (
            olympiad.start_datetime + timedelta(minutes=olympiad.duration_minutes)
            if olympiad.start_datetime and olympiad.duration_minutes else None
        )
        if end_time and now > end_time:
            return Response({'detail': 'Olimpiada vaqti tugagan'},
                            status=http_status.HTTP_400_BAD_REQUEST)
        if not user_can_participate_in_event(request.user, olympiad):
            detail = (
                "Musobaqaga qatnashish uchun shu o'quv markaz tasdig'i kerak"
                if olympiad.event_type == Olympiad.EVENT_TYPE_COMPETITION
                else 'Forbidden'
            )
            return Response(
                {'detail': detail},
                status=http_status.HTTP_403_FORBIDDEN,
            )

        session = (
            TestSession.objects
            .select_for_update()
            .filter(user=request.user, olympiad=olympiad)
            .first()
        )
        if not session:
            return Response(
                {'detail': "Avval test savollarini boshlang"},
                status=http_status.HTTP_400_BAD_REQUEST,
            )
        if session.status == TestSession.STATUS_DISQUALIFIED:
            return Response(
                {'detail': "Siz cheating qildingiz. Olimpiada yakunlandi."},
                status=http_status.HTTP_403_FORBIDDEN,
            )
        # 60 soniyalik grace period: timer tugaganda yuborilgan submit sekin
        # tarmoqda kechikib kelsa ham javoblar yo'qolmasin.
        if session_is_expired(session, olympiad, grace_seconds=SUBMIT_GRACE_SECONDS):
            return Response({'detail': 'Test vaqti tugagan'},
                            status=http_status.HTTP_400_BAD_REQUEST)

        # IT olimpiadasi til cheklovi: allowed_languages to'ldirilgan bo'lsa,
        # yuborilgan har bir kod javobi shu tillardan biri bo'lishi shart.
        # Frontend ham tekshiradi, lekin server avtoritar.
        code_answers = serializer.validated_data.get('code_answers') or {}
        allowed_langs = [
            str(lang).strip().lower()
            for lang in (olympiad.allowed_languages or [])
            if str(lang).strip()
        ]
        if allowed_langs and code_answers:
            for payload in code_answers.values():
                lang = str((payload or {}).get('language') or '').strip().lower()
                # Bo'sh til — savolning default tilidan foydalaniladi, taqiqlamaymiz.
                if lang and lang not in allowed_langs:
                    return Response(
                        {
                            'detail': (
                                "Bu olimpiadada faqat "
                                f"{', '.join(allowed_langs)} ishlatiladi"
                            ),
                        },
                        status=http_status.HTTP_400_BAD_REQUEST,
                    )

        answers = serializer.validated_data.get('answers', {}) or {}
        scored = score_session_answers(session, olympiad, answers)
        total = scored['total']
        correct = scored['correct']
        wrong = scored['wrong']
        max_possible = scored['max_possible']
        score = scored['score']
        time_spent = max(0, int((timezone.now() - session.started_at).total_seconds()))
        if olympiad.duration_minutes:
            time_spent = min(time_spent, olympiad.duration_minutes * 60)

        # Savepoint (nested atomic) — outer transaction'ni abort qilmasdan
        # IntegrityError'ni catch qilamiz. Aks holda Django outer block'da
        # TransactionManagementError otadi va keyingi DB so'rovlari ishlamaydi.
        duplicate_existing = None
        try:
            with transaction.atomic():
                attempt = TestAttempt.objects.create(
                    user=request.user,
                    olympiad=olympiad,
                    answers=answers,
                    score=score,
                    correct_count=correct,
                    wrong_count=wrong,
                    total_questions=total,
                    time_spent=time_spent,
                    rank=None,
                )
        except IntegrityError:
            # Race condition: bir vaqtda ikkita so'rov yuborilsa yoki bir
            # foydalanuvchi tezda ikki marta bossa unique constraint
            # (user, olympiad) ishga tushadi. Mavjud attempt'ni qaytaramiz.
            duplicate_existing = TestAttempt.objects.filter(
                user=request.user, olympiad=olympiad,
            ).first()
            attempt = None

        if duplicate_existing is not None:
            # Sessionni ham COMPLETED ga o'tkazish kerak — aks holda
            # session_is_expired tekshiruvi keyingi marta bo'sh natija
            # qaytaradi va frontend "Test vaqti tugagan" xato ko'rsatardi.
            if session.status != TestSession.STATUS_COMPLETED:
                session.status = TestSession.STATUS_COMPLETED
                session.save(update_fields=['status'])
            data = TestAttemptSerializer(duplicate_existing).data
            data['max_score'] = max_possible
            data['blank_count'] = scored.get('blank', 0)
            data['answered_count'] = scored.get('answered', 0)
            data['detail'] = 'Siz allaqachon topshirgansiz'
            return Response(data, status=http_status.HTTP_200_OK)
        if attempt is None:
            return Response(
                {'detail': "Siz bu olimpiadaga allaqachon qatnashgansiz"},
                status=http_status.HTTP_409_CONFLICT,
            )

        # Re-rank butun jadvalni emas, faqat shu foydalanuvchining
        # rank'ini bitta COUNT query orqali hisoblaymiz va shu attempt'ga
        # yozamiz. Bu submit DB yukini minimal saqlaydi (boshqa
        # attempts'lar update qilinmaydi), lekin sertifikat endpoint'da
        # `rank==1` tekshiruvi to'g'ri ishlashi uchun saqlangan rank
        # qiymati zarur. Boshqa qatnashchilarning rank'lari leaderboard
        # `order_by` orqali jonli tartiblanadi yoki olimpiada yopilganda
        # `recompute_olympiad_ranks` orqali yangilanadi.
        better_count_for_rank = TestAttempt.objects.filter(
            olympiad=olympiad,
            disqualified=False,
        ).filter(
            Q(score__gt=score)
            | Q(score=score, time_spent__lt=time_spent),
        ).exclude(pk=attempt.pk).count()
        attempt.rank = better_count_for_rank + 1
        attempt.save(update_fields=['rank'])

        # Coins va streak'ni atomic blok ichida locked user ustida yangilaymiz.
        # Avval `request.user` (stale, lock qilinmagan) ustida yangilanardi va
        # bir foydalanuvchi parallel ravishda submit + reward redeem qilsa
        # coins qiymati ustiga yozilib (lost update) yo'qolardi.
        from django.contrib.auth import get_user_model
        User = get_user_model()
        try:
            locked_user = User.objects.select_for_update().get(pk=request.user.pk)
            earned_coins = (correct * 10) + 20
            locked_user.coins = (locked_user.coins or 0) + earned_coins
            locked_user.save(update_fields=['coins'])
            locked_user.update_streak()
            # Joriy request.user ob'ektini ham yangilab qo'yamiz — quyida
            # streak_count javobga qo'shiladi.
            request.user.coins = locked_user.coins
            request.user.streak_count = locked_user.streak_count
        except Exception:
            import logging
            logging.getLogger(__name__).exception(
                'coins/streak update failed for user=%s attempt=%s',
                request.user.pk, attempt.id,
            )

        # O2: Kunlik maqsadni javob berilgan savollar soni bilan oldinga
        # suramiz (maqsad bajarilsa +50 coin). Hech qachon submit'ni buzmaydi.
        daily_goal_achieved = False
        try:
            from accounts.daily_goal import record_progress
            answered_for_goal = scored.get('answered', correct + wrong)
            daily_goal_achieved = record_progress(
                request.user, answered_for_goal,
                locked_user=locals().get('locked_user'),
            )
        except Exception:
            import logging
            logging.getLogger(__name__).exception(
                'daily goal update failed for user=%s attempt=%s',
                request.user.pk, attempt.id,
            )

        session.status = TestSession.STATUS_COMPLETED
        session.save(update_fields=['status'])

        # IT (kod) javoblarini saqlaymiz va AI baholashni background'da ishga
        # tushuramiz. Oddiy MCQ olimpiadalarda code_answers bo'sh — bu blok
        # hech narsa qilmaydi. Submit'ni hech qachon buzmaydi.
        try:
            _save_code_submissions(
                attempt, olympiad,
                serializer.validated_data.get('code_answers') or {},
            )
        except Exception:
            import logging
            logging.getLogger(__name__).exception(
                'code submission save failed for attempt=%s', attempt.id,
            )

        # O5: Yutuq (milestone) tekshiruvi — submit'dan keyin. request.user
        # yuqorida streak_count bilan yangilangan, shu sababli streak
        # milestone'lari to'g'ri hisoblanadi. Hech qachon exception otmaydi.
        try:
            from accounts.achievements import check_achievements
            check_achievements(request.user, attempt)
        except Exception:
            import logging
            logging.getLogger(__name__).exception(
                'achievement check failed for attempt=%s', attempt.id,
            )

        # Adaptiv daraja (ELO'ga o'xshash): natija olimpiada fani bo'yicha
        # foydalanuvchining darajasini yuqori/past yo'naltiradi. score 0–100.
        # Faqat foydalanuvchi shu fanga onboarding'da daraja belgilagan bo'lsa
        # ishlaydi. Hech qachon submit'ni buzmaydi.
        try:
            subject = getattr(olympiad, 'subject', None) or ''
            user = request.user
            if subject and user.subject_levels and subject in user.subject_levels:
                if score >= 70:
                    user.update_subject_level(subject, 'up')
                elif score < 40:
                    user.update_subject_level(subject, 'down')
                # 40–69 — neytral, daraja o'zgarmaydi.
        except Exception:
            import logging
            logging.getLogger(__name__).exception(
                'subject level update failed for attempt=%s', attempt.id,
            )

        # Yangi attempt qo'shildi — bashorat (predictions) cache'i endi
        # eskirgan (o'rtacha ball va fan kesimi o'zgardi). Keyingi
        # /me/predictions/ so'rovida qayta hisoblanishi uchun bekor qilamiz.
        try:
            from accounts.utils import invalidate_user_predictions_cache
            invalidate_user_predictions_cache(request.user.id)
        except Exception:
            pass

        # O4: Premium o'quvchi uchun avtomatik AI tahlil. Submit latency'ni
        # bloklamaslik uchun: shu yerda pending yozuv yaratamiz, AI call'ni
        # alohida daemon thread'da bajaramiz (Gemini 45s gacha kutishi
        # mumkin — bu submit'ni bloklamasligi shart). Endpoint tayyor
        # bo'lmaguncha {status: "pending"} qaytaradi.
        from accounts.utils import is_user_premium
        if is_user_premium(request.user):
            try:
                _trigger_attempt_ai_analysis(attempt, olympiad, answers)
            except Exception:
                import logging
                logging.getLogger(__name__).exception(
                    'AI analysis trigger failed for attempt=%s', attempt.id,
                )

        # Center rating recompute ham submit ichidan olib tashlandi —
        # bu N ta attempt bo'yicha katta AVG aggregate qilardi va submit
        # latency'ni oshirardi. Funksiyaning o'zi mavjud, kerak bo'lsa
        # cron yoki manager dashboard'ida chaqiriladi.

        # Ota-onalarga Telegram xabari. Sinxron API call'ni request thread'ini
        # bloklamasligi uchun Celery orqali asinxron yuboramiz.
        try:
            from .tasks import send_attempt_result_to_parents_task
            send_attempt_result_to_parents_task.delay(attempt.id)
        except Exception:
            import logging
            logging.getLogger(__name__).exception(
                'failed to queue parent notification task'
            )

        data = TestAttemptSerializer(attempt).data
        data['max_score'] = max_possible
        # Yangi semantika: blank (javob berilmagan) va answered alohida
        # qaytariladi. Frontend "Sizning natijangiz" sahifasida 4 ta
        # sonni alohida ko'rsatishi mumkin: to'g'ri / noto'g'ri / bo'sh.
        data['blank_count'] = scored.get('blank', 0)
        data['answered_count'] = scored.get('answered', 0)
        # request.user.streak_count yuqorida locked_user'dan yangilangan —
        # ishonchli, lock'langan qiymatni qaytaramiz.
        data['streak_count'] = request.user.streak_count
        # O2: kunlik maqsad shu submit'da bajarildimi (frontend bonus
        # animatsiyasi uchun).
        data['daily_goal_achieved'] = daily_goal_achieved
        # Rank submit paytida yuqorida `attempt.rank` ga yozildi — bu
        # qiymatni ham `position` field sifatida qaytaramiz (frontend
        # eski kod position'ga tayanishi mumkin).
        data['position'] = attempt.rank or (better_count_for_rank + 1)
        # Backward-compat: agar TestAttemptSerializer rank'ni qaytarmasa
        # (eski serializer), shu yerda qo'shamiz.
        if data.get('rank') is None:
            data['rank'] = attempt.rank or (better_count_for_rank + 1)
        return Response(data, status=http_status.HTTP_201_CREATED)


submit_attempt.cls.throttle_scope = 'submit'


class ReportCheatingView(APIView):
    """POST /api/attempts/cheating/ — disqualify current user's test session.

    Throttle: foydalanuvchi bir daqiqada 5 martadan ortiq cheating signal
    yubora olmaydi — aks holda olimpiada paytida frontend bug yoki yomon
    niyatli skript orqali DB'ga bosim kelishi mumkin. `throttle_scope`
    DEFAULT_THROTTLE_RATES['cheating'] = '5/min' ga ulanadi.

    Eslatma: avval FBV (`@api_view`) edi va throttle scope `report_cheating.cls.
    throttle_scope = 'cheating'` orqali (DRF ichki `.cls` atributiga) o'rnatilardi.
    Standart CBV'da scope shunchaki klass atributi — `.cls` hiylasi kerak emas.
    """
    permission_classes = [IsAuthenticated]
    throttle_classes = [ScopedRateThrottle]
    throttle_scope = 'cheating'

    def post(self, request):
        olympiad_id = request.data.get('olympiad')
        reason = str(request.data.get('reason') or 'test_window_left')[:120]
        if not olympiad_id:
            return Response({'detail': 'olympiad majburiy'}, status=http_status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            olympiad = get_object_or_404(
                Olympiad.objects.select_for_update().select_related('center', 'center__owner')
                # _build_attempt_mistakes va scoring `olympiad.questions.all()` ni
                # aylanadi — savollarni oldindan yuklab N+1 so'rovlarni oldini olamiz.
                .prefetch_related('questions'),
                pk=olympiad_id,
            )
            if not user_can_participate_in_event(request.user, olympiad):
                return Response({'detail': 'Forbidden'}, status=http_status.HTTP_403_FORBIDDEN)
            if TestAttempt.objects.filter(user=request.user, olympiad=olympiad).exists():
                return Response({'disqualified': False, 'detail': 'Attempt already submitted'})
            session = (
                TestSession.objects
                .select_for_update()
                .filter(user=request.user, olympiad=olympiad)
                .first()
            )
            if not session:
                return Response({'detail': "Test session topilmadi"}, status=http_status.HTTP_400_BAD_REQUEST)
            if session.status == TestSession.STATUS_COMPLETED:
                return Response({'disqualified': False, 'detail': 'Attempt already submitted'})
            notify = session.status != TestSession.STATUS_DISQUALIFIED
            session.status = TestSession.STATUS_DISQUALIFIED
            session.disqualified_at = session.disqualified_at or timezone.now()
            session.cheating_reason = session.cheating_reason or reason
            session.save(update_fields=['status', 'disqualified_at', 'cheating_reason'])

            # Diskvalifikatsiya bo'lgan student uchun ham attempt yaratamiz —
            # aks holda na leaderboard'da, na manager paneli statistikasida
            # ko'rinmasdi. score=0, disqualified=True bilan iz qoldiramiz.
            # Session boshlanganidan hozirgacha bo'lgan vaqtni time_spent qilamiz.
            time_spent = max(0, int(
                (timezone.now() - session.started_at).total_seconds()
            )) if session.started_at else 0
            if olympiad.duration_minutes:
                time_spent = min(time_spent, olympiad.duration_minutes * 60)
            try:
                TestAttempt.objects.create(
                    user=request.user,
                    olympiad=olympiad,
                    answers={},
                    score=0,
                    correct_count=0,
                    wrong_count=0,
                    total_questions=0,
                    time_spent=time_spent,
                    rank=None,
                    disqualified=True,
                )
            except IntegrityError:
                # Race: bir vaqtda submit bilan kelishi mumkin. E'tibor bermaymiz.
                pass

        if notify:
            try:
                from notifications.services import send_cheating_detected_notification

                send_cheating_detected_notification(request.user, olympiad, olympiad.center, reason)
            except Exception:
                import logging
                logging.getLogger(__name__).exception('cheating notification failed')
        return Response({
            'disqualified': True,
            'detail': "Siz cheating qildingiz. Olimpiada yakunlandi.",
        })


# URL routing FBV-shaklidagi callable kutadi — CBV'ni `.as_view()` orqali
# beramiz, shunda urls.py'dagi `views.report_cheating` o'zgarmasdan ishlaydi.
report_cheating = ReportCheatingView.as_view()


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def attempt_detail(request, attempt_id):
    """GET /api/attempts/{attempt_id}/ — bitta attempt natijasi.

    Leaderboard "Ko'rish" tugmasi va shunga o'xshash sahifalar uchun.
    Ruxsatlar:
      - Attempt egasi (foydalanuvchining o'zi)
      - Platform admin
      - Olympiad markaziga tegishli owner/manager
    Boshqa hollarda 403 qaytariladi.
    """
    from django.db.models import Prefetch
    from questions.models import Question
    attempt = get_object_or_404(
        TestAttempt.objects
        .filter(olympiad__is_deleted=False)
        .select_related('user', 'olympiad', 'olympiad__center')
        # Quyida savollar `id` bo'yicha tartiblangan holda aylanadi —
        # savollarni oldindan yuklab N+1 (har savol uchun alohida so'rov)
        # o'rniga 1 ta so'rov. Tartiblashni Prefetch queryset ichida
        # qilamiz: aks holda `.order_by('id')` prefetch cache'ni buzib
        # qo'shimcha DB so'rovi otardi.
        .prefetch_related(
            Prefetch(
                'olympiad__questions',
                queryset=Question.objects.order_by('id'),
            ),
            # Kod (IT) javoblari va essay baholarini ham oldindan yuklaymiz —
            # quyida savol bo'yicha map qilinadi (N+1 emas, 1 ta so'rov).
            'code_submissions',
            'essay_grades',
        ),
        pk=attempt_id,
    )
    is_owner = attempt.user_id == request.user.id
    is_admin = request.user.is_platform_admin
    can_view = is_owner or is_admin
    if not can_view and attempt.olympiad.center_id:
        can_view = (
            attempt.olympiad.center.owner_id == request.user.id
            or CenterMembership.objects.filter(
                user=request.user,
                center_id=attempt.olympiad.center_id,
                role__in=[CenterMembership.ROLE_MANAGER, CenterMembership.ROLE_OWNER],
                status=CenterMembership.STATUS_APPROVED,
            ).exists()
        )
    if not can_view:
        return Response({'detail': 'Forbidden'}, status=http_status.HTTP_403_FORBIDDEN)

    data = TestAttemptSerializer(attempt).data
    # Olimpiada ma'lumotini ham qo'shamiz — frontend bittagina so'rov bilan
    # to'liq sahifani chizishi mumkin.
    olympiad = attempt.olympiad
    data['olympiad_detail'] = {
        'id': olympiad.id,
        'title': olympiad.title,
        'subject': olympiad.subject,
        'event_type': olympiad.event_type,
        'test_level': olympiad.test_level,
        'test_type': olympiad.test_type,
        'duration_minutes': olympiad.duration_minutes,
        'start_datetime': olympiad.start_datetime.isoformat() if olympiad.start_datetime else None,
        'center_id': olympiad.center_id,
        'center_name': olympiad.center.name if olympiad.center_id else '',
    }
    # Review mode: faqat attempt egasi (yoki manager/admin) savollarni
    # ko'rishi mumkin. Bu javoblarni tahlil qilish uchun ishlatiladi —
    # foydalanuvchi qaysi savolda xato qilganini ko'rsata oladi.
    if is_owner or is_admin or (not is_owner and can_view):
        questions_review = []
        # answers dict kalitlari string yoki integer bo'lishi mumkin.
        answers = attempt.answers or {}
        # Kod (IT) javoblari — yuqorida `prefetch_related('code_submissions')`
        # bilan oldindan yuklangan; `.all()` orqali prefetch cache'dan o'qiymiz
        # (qo'shimcha DB so'rovi otmaymiz). `.filter(...)` ishlatsak prefetch
        # e'tiborsiz qolib N+1 qaytarди.
        code_subs = {
            cs.question_id: cs
            for cs in attempt.code_submissions.all()
        }
        # Essay baholari (qo'lda) — savol bo'yicha map. Baholangan essay
        # "tekshirilmoqda" o'rniga ball + izoh bilan ko'rsatiladi.
        essay_grades = {
            g.question_id: g
            for g in attempt.essay_grades.all()
        }
        # Prefetch `id` bo'yicha tartiblangan — bu yerda `.order_by('id')`
        # qo'ymaymiz, aks holda prefetch cache buziladi va yangi DB so'rovi
        # otiladi.
        for q in olympiad.questions.all():
            q_type = getattr(q, 'question_type', 'mcq') or 'mcq'
            if q_type == 'code':
                cs = code_subs.get(q.id)
                questions_review.append({
                    'id': q.id,
                    'text': q.text,
                    'options': [],
                    'question_type': 'code',
                    'programming_language': getattr(q, 'programming_language', '') or '',
                    'difficulty': q.difficulty,
                    'score': q.score,
                    'subject': q.subject,
                    'submitted_code': cs.submitted_code if cs else '',
                    'code_language': cs.code_language if cs else '',
                    'ai_code_review': cs.ai_code_review if cs else '',
                    'ai_code_score': cs.ai_code_score if cs else None,
                })
                continue
            raw_chosen = answers.get(str(q.id))
            if raw_chosen is None:
                raw_chosen = answers.get(q.id)
            chosen_val = _extract_review_chosen(raw_chosen, q_type)

            # Yangi savol turlari (multiple_select/fill_blank/fill_blanks/
            # yes_no/essay) — grade_answer orqali baholanadi. Essay
            # (RESULT_PENDING) avtomatik baholanmaydi: alohida "tekshirilmoqda"
            # holatida ko'rsatiladi (is_correct=None), 0 ball/noto'g'ri emas.
            review_item = {
                'id': q.id,
                'text': q.text,
                'options': q.options or [],
                'question_type': q_type,
                'difficulty': q.difficulty,
                'score': q.score,
                'subject': q.subject,
            }
            if q_type in ('mcq', 'yes_no'):
                # Mavjud xatti-harakatni saqlaymiz: chosen_idx ni qaytaramiz va
                # correct_answer indeksi bilan solishtiramiz (frontend
                # variantlarni shu indeks bo'yicha belgilaydi).
                try:
                    chosen_idx = int(chosen_val) if chosen_val is not None else None
                except (TypeError, ValueError):
                    chosen_idx = None
                review_item['correct_answer'] = q.correct_answer
                review_item['chosen_answer'] = chosen_idx
                review_item['is_correct'] = (
                    chosen_idx is not None and chosen_idx == q.correct_answer
                )
            elif q_type == 'essay':
                # Qo'lda baholanadi — natija sahifasida alohida ko'rsatiladi.
                # Ustoz baho qo'ygan bo'lsa ball + izoh qaytariladi va
                # pending_review=False bo'ladi.
                review_item['chosen_answer'] = chosen_val
                grade = essay_grades.get(q.id)
                if grade is not None:
                    review_item['pending_review'] = False
                    review_item['essay_score'] = grade.score
                    review_item['essay_feedback'] = grade.feedback or ''
                    review_item['is_correct'] = grade.score >= q.score
                else:
                    review_item['pending_review'] = True
                    review_item['is_correct'] = None
            else:
                # multiple_select / fill_blank / fill_blanks — server baholaydi.
                # Frontend to'g'ri javobni ko'rsata olishi uchun korrekt
                # qiymatni ham qaytaramiz: multiple_select → to'g'ri option
                # indekslari (correct_answer_set), fill_blank/fill_blanks →
                # matn yoki {"1": "..."} dict (correct_text).
                from questions.grading import _parse_correct_text
                result = grade_answer(q, chosen_val)
                review_item['chosen_answer'] = chosen_val
                review_item['is_correct'] = (result == RESULT_CORRECT)
                if q_type == 'multiple_select':
                    correct_raw = _parse_correct_text(getattr(q, 'correct_text', ''))
                    if isinstance(correct_raw, (list, tuple)):
                        try:
                            review_item['correct_answer_set'] = [int(x) for x in correct_raw]
                        except (TypeError, ValueError):
                            review_item['correct_answer_set'] = []
                else:
                    # fill_blank / fill_blanks — to'g'ri matn(lar).
                    review_item['correct_text'] = _parse_correct_text(
                        getattr(q, 'correct_text', ''),
                    )
            questions_review.append(review_item)
        data['questions_review'] = questions_review
    return Response(data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def attempt_ai_analysis(request, attempt_id):
    """GET /api/attempts/<id>/ai-analysis/ — saqlangan AI tahlilini qaytaradi (O4).

    Faqat attempt egasi (yoki admin). Tahlil hali tayyor bo'lmasa
    {status: "pending"}, tayyor bo'lsa {status: "ready", analysis: "..."}.
    Premium bo'lmagan foydalanuvchi uchun 403.
    """
    from .models import AttemptAIAnalysis

    attempt = get_object_or_404(
        TestAttempt.objects.select_related('user', 'olympiad'),
        pk=attempt_id,
    )
    is_owner = attempt.user_id == request.user.id
    if not (is_owner or request.user.is_platform_admin):
        return Response({'detail': 'Forbidden'}, status=http_status.HTTP_403_FORBIDDEN)
    # Real-time premium tekshiruvi: flag + aktiv obuna muddati (60s cache).
    from accounts.utils import is_user_premium
    owner_is_premium = is_user_premium(request.user) if is_owner else False
    if is_owner and not owner_is_premium:
        return Response(
            {
                'detail': "AI tahlil premium o'quvchilar uchun.",
                'upgrade_required': True,
            },
            status=http_status.HTTP_403_FORBIDDEN,
        )

    analysis = AttemptAIAnalysis.objects.filter(attempt=attempt).first()
    if not analysis:
        # Hali umuman trigger qilinmagan bo'lsa (masalan, eski attempt yoki
        # submit paytida premium bo'lmagan). Egasi premium bo'lsa shu yerda
        # lazy ravishda boshlаymiz.
        if is_owner and owner_is_premium:
            try:
                _trigger_attempt_ai_analysis(
                    attempt, attempt.olympiad, attempt.answers or {},
                )
            except Exception:
                import logging
                logging.getLogger(__name__).exception(
                    'lazy AI analysis trigger failed for attempt=%s', attempt.id,
                )
        return Response({'status': 'pending'})
    if analysis.status == AttemptAIAnalysis.STATUS_READY:
        return Response({'status': 'ready', 'analysis': analysis.analysis_text})
    if analysis.status == AttemptAIAnalysis.STATUS_FAILED:
        return Response({'status': 'failed', 'analysis': analysis.analysis_text or ''})
    return Response({'status': 'pending'})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def my_results(request):
    """GET /api/results/me/ — current user's attempt history.

    Pagination qo'llab-quvvatlanadi: `?page=`, `?page_size=` (default 50,
    max 200). Eski klientlar ham ishlashda davom etadi — birinchi page
    avtomatik qaytariladi. Avval qattiq `[:200]` cheklov edi va keyingi
    attempt'lar yashirin bo'lib qolardi.
    """
    # Soft-delete qilingan olimpiadalar foydalanuvchining tarixida
    # ko'rinmasin — manager o'chirgan tadbir egasi nuqtai nazaridan
    # ham yo'q deb hisoblanadi.
    qs = (
        TestAttempt.objects
        .filter(user=request.user, olympiad__is_deleted=False)
        .select_related('olympiad')
    )
    try:
        page = int(request.query_params.get('page') or 1)
    except (TypeError, ValueError):
        page = 1
    page = max(1, page)
    try:
        page_size = int(request.query_params.get('page_size') or 50)
    except (TypeError, ValueError):
        page_size = 50
    page_size = max(1, min(page_size, 200))
    total = qs.count()
    offset = (page - 1) * page_size
    rows = qs[offset:offset + page_size]
    data = TestAttemptSerializer(rows, many=True).data
    # Backward-compat: agar klient `?page=` yubormagan va `?page_size=` ham
    # yubormagan bo'lsa, javobni list ko'rinishida qaytaramiz (eski
    # StudentDashboard kodi shu formatga tayanadi). Aks holda standart
    # pagination dict.
    if not request.query_params.get('page') and not request.query_params.get('page_size'):
        return Response(data)
    return Response({
        'results': data,
        'pagination': {
            'page': page,
            'page_size': page_size,
            'total': total,
            'has_next': offset + len(data) < total,
        },
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def my_stats(request):
    """GET /api/results/me/stats/ — aggregated per-subject stats.

    Returns:
      {
        total_attempts, average_score, best_rank,
        subjects: [{subject, attempts, average_score}, ...]
      }
    """
    qs = TestAttempt.objects.filter(user=request.user)

    # Umumiy statistikani DB darajasida hisoblaymiz (attempt'larni xotiraga
    # yuklamasdan). best_rank uchun faqat NULL bo'lmagan ranklar hisobga olinadi.
    overall = qs.aggregate(
        total=Count('id'),
        avg=Avg('score'),
        best_rank=Min('rank'),
    )
    total = overall['total'] or 0
    if total == 0:
        return Response({
            'total_attempts': 0,
            'average_score': 0,
            'best_rank': None,
            'subjects': [],
        })
    avg = round(overall['avg'] or 0, 1)
    best_rank = overall['best_rank']

    # Fan kesimida o'rtacha ball — DB'da GROUP BY orqali.
    subjects = []
    subject_rows = (
        qs.values('olympiad__subject')
        .annotate(attempts=Count('id'), avg_score=Avg('score'))
    )
    for row in subject_rows:
        subjects.append({
            'subject': row['olympiad__subject'] or '—',
            'attempts': row['attempts'],
            'average_score': round(row['avg_score'] or 0, 1),
        })
    subjects.sort(key=lambda x: -x['average_score'])
    return Response({
        'total_attempts': total,
        'average_score': avg,
        'best_rank': best_rank,
        'subjects': subjects,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_certificate(request, attempt_id):
    """GET /api/certificates/{attempt_id}/download/ — PNG sertifikat.

    Sertifikat alohida modelda saqlanmaydi: TestAttempt'dan har safar
    on-the-fly generatsiya qilinadi. Faqat 1-o'rin egasi sertifikat ola
    oladi — boshqa ishtirokchilar uchun 403.
    """
    attempt = get_object_or_404(
        TestAttempt.objects.select_related('user', 'olympiad', 'olympiad__center'),
        pk=attempt_id,
    )
    # Faqat o'z attempti yoki center managerlari/admin
    is_owner = attempt.user_id == request.user.id
    is_admin = request.user.is_platform_admin
    can_view = is_owner or is_admin
    if not can_view and attempt.olympiad.center_id:
        can_view = CenterMembership.objects.filter(
            user=request.user,
            center_id=attempt.olympiad.center_id,
            role__in=[CenterMembership.ROLE_MANAGER, CenterMembership.ROLE_OWNER],
            status=CenterMembership.STATUS_APPROVED,
        ).exists() or attempt.olympiad.center.owner_id == request.user.id
    if not can_view:
        return Response({'detail': 'Forbidden'}, status=http_status.HTTP_403_FORBIDDEN)
    # Eski attempt'larda `rank` NULL bo'lishi mumkin (submit paytida
    # rank yangilanmagan eski siyosat). Bunday hollarda jonli pozitsiyani
    # bitta COUNT query orqali hisoblab, DB'da rank field'ni yangilab
    # qo'yamiz — keyingi safarda qayta hisoblanmaydi.
    if attempt.rank is None:
        live_better = TestAttempt.objects.filter(
            olympiad=attempt.olympiad,
            disqualified=False,
        ).filter(
            Q(score__gt=attempt.score)
            | Q(score=attempt.score, time_spent__lt=attempt.time_spent),
        ).exclude(pk=attempt.pk).count()
        attempt.rank = live_better + 1
        attempt.save(update_fields=['rank'])
    # Faqat 1-o'rin egasi sertifikat ola oladi. Avval har qanday rank
    # uchun ruxsat berilardi — bu sertifikatni hammaga tarqatib yuborardi.
    if attempt.rank != 1:
        return Response(
            {'detail': "Sertifikat faqat 1-o'rin egasiga beriladi"},
            status=http_status.HTTP_403_FORBIDDEN,
        )
    # Eski attempt'larda certificate_uuid NULL bo'lishi mumkin (migratsiya
    # data fill faqat eski qatorlarni qoplaydi, lekin imkoniyat uchun bu
    # yerda ham lazy to'ldiramiz). Sertifikat PNG'sida verify URL shu UUID
    # orqali ko'rsatiladi.
    if not attempt.certificate_uuid:
        import uuid as _uuid
        attempt.certificate_uuid = _uuid.uuid4()
        attempt.save(update_fields=['certificate_uuid'])
    from django.core.cache import cache
    cache_key = f"certificate:attempt:{attempt.id}"
    png_bytes = cache.get(cache_key)
    if not png_bytes:
        try:
            png_bytes = render_certificate_png(attempt)
            cache.set(cache_key, png_bytes, timeout=604800)  # 7 days
        except Exception as exc:
            import logging
            logging.getLogger(__name__).exception('certificate render failed: %s', exc)
            return Response(
                {'detail': "Sertifikatni yaratib bo'lmadi"},
                status=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
    response = HttpResponse(png_bytes, content_type='image/png')
    safe_title = ''.join(ch for ch in attempt.olympiad.title if ch.isalnum() or ch in (' ', '_', '-'))[:60].strip() or 'certificate'
    response['Content-Disposition'] = f'attachment; filename="olympy-{safe_title}-{attempt.id}.png"'
    return response


@api_view(['GET'])
@permission_classes([AllowAny])
def certificate_verify(request, cert_uuid):
    """GET /api/certificates/verify/{uuid}/ — sertifikat haqiqiyligini tekshirish.

    PUBLIC (auth shart emas) — sertifikatdagi QR/URL ochilganda kim bo'lishidan
    qat'i nazar tekshirishi mumkin. Javob `reason` orqali holatni aniq
    ajratadi, shunda foydalanuvchi nima xato bo'lganini tushunadi:
      - UUID topilmadi → {valid: false, reason: "not_found"} 404
      - UUID bor, lekin sertifikat berilmagan (disqualified yoki rank!=1)
        → {valid: false, reason: "not_awarded"} 404
      - To'g'ri sertifikat → {valid: true, reason: "ok", ...} 200

    Faqat 1-o'rin egasiga sertifikat beriladigani uchun (download_certificate
    bilan bir xil siyosat) — rank!=1 bo'lgan attempt UUID'i ham haqiqiy
    sertifikat hisoblanmaydi.
    """
    attempt = (
        TestAttempt.objects
        .filter(certificate_uuid=cert_uuid, olympiad__is_deleted=False)
        .select_related('user', 'olympiad', 'olympiad__center')
        .first()
    )
    if not attempt:
        return Response(
            {'valid': False, 'reason': 'not_found'},
            status=http_status.HTTP_404_NOT_FOUND,
        )
    if attempt.disqualified or attempt.rank != 1:
        return Response(
            {'valid': False, 'reason': 'not_awarded'},
            status=http_status.HTTP_404_NOT_FOUND,
        )

    olympiad = attempt.olympiad
    center = olympiad.center if olympiad.center_id else None
    student_name = (attempt.user.full_name or 'Foydalanuvchi').strip()
    return Response({
        'valid': True,
        'reason': 'ok',
        'student_name': student_name,
        'olympiad_name': olympiad.title,
        'score': attempt.score,
        'date': attempt.submitted_at.strftime('%d.%m.%Y') if attempt.submitted_at else '',
        'center_name': center.name if center else '',
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def manager_stats(request):
    """GET /api/manager/stats/?center=<id> — center natijalarini umumlashtiradi.

    Manager/Owner/Admin uchun: o'rtacha ball, eng yuqori ball, qatnashuvchilar
    soni va tadbirlar bo'yicha breakdown. Hardcoded mock o'rniga real ma'lumot.
    """
    center_id = request.query_params.get('center')
    if not center_id:
        # Auto-pick: foydalanuvchining birinchi manager/owner centeri
        membership = (
            CenterMembership.objects
            .filter(
                user=request.user,
                role__in=[
                    CenterMembership.ROLE_MANAGER,
                    CenterMembership.ROLE_OWNER,
                ],
                status=CenterMembership.STATUS_APPROVED,
            )
            .order_by('-created_at')
            .first()
        )
        if not membership and not request.user.is_platform_admin:
            return Response(
                {'detail': "Center aniqlanmadi. ?center=<id> kiriting"},
                status=http_status.HTTP_400_BAD_REQUEST,
            )
        center_id = membership.center_id if membership else None
    if not center_id:
        return Response(
            {'detail': "Center aniqlanmadi"},
            status=http_status.HTTP_400_BAD_REQUEST,
        )
    center = get_object_or_404(EducationCenter, pk=center_id)
    # Auth check: faqat manager/owner/admin
    is_admin = request.user.is_platform_admin
    is_owner = center.owner_id == request.user.id
    is_manager = CenterMembership.objects.filter(
        user=request.user, center=center,
        role=CenterMembership.ROLE_MANAGER,
        status=CenterMembership.STATUS_APPROVED,
    ).exists()
    if not (is_admin or is_owner or is_manager):
        return Response({'detail': 'Forbidden'}, status=http_status.HTTP_403_FORBIDDEN)

    qs = TestAttempt.objects.filter(olympiad__center=center, olympiad__is_deleted=False)
    # Diskvalifitsiya bo'lganlar agregatga kirmaydi — aks holda o'rtacha
    # ball nohaq pasayardi. Lekin alohida hisob sifatida `disqualified_count`
    # ham qaytariladi.
    qs_valid = qs.filter(disqualified=False)
    agg = qs_valid.aggregate(
        avg=Avg('score'),
        best=Max('score'),
        participants=Count('user', distinct=True),
        total_attempts=Count('id'),
    )
    disqualified_count = qs.filter(disqualified=True).count()
    # Per-event breakdown — finished + active tadbirlar bo'yicha.
    # Avval har bir olimpiada uchun alohida aggregate query ishga tushar
    # (50 ta olimpiada bo'lsa 50 ta SQL). Endi bitta GROUP BY query orqali
    # barcha olimpiadalarning agregati olinadi, keyin Python'da olimpiada
    # meta-ma'lumotlari bilan birlashtiriladi.
    # Y9: pagination — `?page=`, `?page_size=` (default 50, max 200).
    # Jami events soni `events_total` orqali qaytariladi, agregat esa
    # butun markaz bo'yicha hisoblanadi (50 limit bilan emas) — bu
    # foydalanuvchi uchun aniqroq.
    olympiads_full_qs = Olympiad.objects.filter(center=center, is_deleted=False)
    events_total = olympiads_full_qs.count()
    try:
        events_page = int(request.query_params.get('page') or 1)
    except (TypeError, ValueError):
        events_page = 1
    events_page = max(1, events_page)
    try:
        events_page_size = int(request.query_params.get('page_size') or 50)
    except (TypeError, ValueError):
        events_page_size = 50
    events_page_size = max(1, min(events_page_size, 200))
    events_offset = (events_page - 1) * events_page_size
    olympiads_qs = list(
        olympiads_full_qs
        .order_by('-created_at')
        [events_offset:events_offset + events_page_size]
    )
    olympiad_ids = [o.id for o in olympiads_qs]
    per_event_aggs = {}
    if olympiad_ids:
        rows = (
            TestAttempt.objects
            .filter(olympiad_id__in=olympiad_ids)
            .values('olympiad_id')
            .annotate(
                avg=Avg('score'),
                best=Max('score'),
                participants=Count('user', distinct=True),
            )
        )
        per_event_aggs = {row['olympiad_id']: row for row in rows}
    events = []
    for o in olympiads_qs:
        sub_agg = per_event_aggs.get(o.id, {})
        events.append({
            'olympiad_id': o.id,
            'title': o.title,
            'subject': o.subject,
            'status': o.status,
            'event_type': o.event_type,
            'average_score': round(sub_agg.get('avg') or 0, 1),
            'best_score': sub_agg.get('best') or 0,
            'participants': sub_agg.get('participants') or 0,
        })
    return Response({
        'center_id': center.id,
        'center_name': center.name,
        'average_score': round(agg['avg'] or 0, 1),
        'best_score': agg['best'] or 0,
        'participants': agg['participants'] or 0,
        'total_attempts': agg['total_attempts'] or 0,
        'disqualified_count': disqualified_count,
        'events': events,
        'events_total': events_total,
        'events_page': events_page,
        'events_page_size': events_page_size,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def question_difficulty_stats(request):
    """GET /api/manager/question-difficulty-stats/?center=<id>

    Markaz savollar bankidagi savollarning qiyinlik bo'yicha taqsimoti va
    har bir qiyinlik darajasi uchun o'rtacha to'g'rilik foizi.
    Avg correct rate hisoblash: har bir savol uchun qatnashganlar orasidan
    nechta to'g'ri javob — yig'indi olib o'rtacha qiymat chiqaramiz.
    """
    from questions.models import Question

    center_id = request.query_params.get('center')
    if not center_id:
        return Response(
            {'detail': 'center query parametri majburiy'},
            status=http_status.HTTP_400_BAD_REQUEST,
        )
    try:
        center_id = int(center_id)
    except (TypeError, ValueError):
        return Response(
            {'detail': "center parametri son bo'lishi kerak"},
            status=http_status.HTTP_400_BAD_REQUEST,
        )
    center = get_object_or_404(EducationCenter, pk=center_id)
    # Auth: faqat manager/owner/teacher/admin shu markaz uchun.
    is_admin = request.user.is_platform_admin
    is_owner = center.owner_id == request.user.id
    is_staff = CenterMembership.objects.filter(
        user=request.user, center=center,
        role__in=[
            CenterMembership.ROLE_MANAGER,
            CenterMembership.ROLE_TEACHER,
            CenterMembership.ROLE_OWNER,
        ],
        status=CenterMembership.STATUS_APPROVED,
    ).exists()
    if not (is_admin or is_owner or is_staff):
        return Response({'detail': 'Forbidden'}, status=http_status.HTTP_403_FORBIDDEN)

    DIFFICULTY_LABELS = {
        'easy': 'Oson',
        'medium': "O'rta",
        'hard': 'Qiyin',
        'beginner': 'Beginner',
        'elementary': 'Elementary',
        'pre-int': 'Pre-Intermediate',
        'int': 'Intermediate',
        'upper-int': 'Upper-Intermediate',
        'advanced': 'Advanced',
    }

    # grade_answer savol turiga qarab correct_text/question_type maydonlaridan
    # ham foydalanadi — ularni .only() ga qo'shamiz, aks holda har bir baholash
    # deferred maydon uchun qo'shimcha DB so'rovi qilib N+1 keltirib chiqaradi.
    questions = list(Question.objects.filter(center_id=center_id).only(
        'id', 'difficulty', 'correct_answer', 'correct_text', 'question_type',
    ))
    total = len(questions)
    if total == 0:
        return Response({'total_questions': 0, 'by_difficulty': []})

    # Markaz tegishli barcha attempts'larni olib, savol bo'yicha
    # to'g'ri/jami javob hisoblaymiz. Bu markaz olimpiadalarida qatnashgan
    # attempts dan kelib chiqadi. Diskvalifikatsiyalar chiqarib tashlanadi.
    # `list(...)` o'rniga `.iterator()` — barcha answers JSON'larni bir vaqtda
    # xotiraga yuklamaymiz, balki oqim (stream) ko'rinishida birma-bir
    # qayta ishlaymiz. Markazda minglab attempt bo'lsa bu xotirani tejaydi.
    attempts = (
        TestAttempt.objects
        .filter(olympiad__center_id=center_id, disqualified=False)
        .values_list('answers', flat=True)
        .iterator(chunk_size=500)
    )

    # Per-question: to'g'ri va jami javob soni.
    qmap = {q.id: q for q in questions}
    correct_count = {q.id: 0 for q in questions}
    answered_count = {q.id: 0 for q in questions}

    for ans in attempts:
        if not isinstance(ans, dict):
            continue
        for k, v in ans.items():
            try:
                qid = int(k)
            except (TypeError, ValueError):
                continue
            q = qmap.get(qid)
            if not q:
                continue
            answered_count[qid] = answered_count.get(qid, 0) + 1
            # `v` savol turiga qarab int (mcq), str (fill_blank), list
            # (multiple_select) yoki dict (fill_blanks) bo'lishi mumkin.
            # `int(v)` yangi turlarda ValueError berardi — grade_answer
            # har bir turni to'g'ri baholaydi.
            if grade_answer(q, v) == RESULT_CORRECT:
                correct_count[qid] = correct_count.get(qid, 0) + 1

    # Difficulty bo'yicha bucket'lar.
    buckets = {}
    for q in questions:
        diff = q.difficulty or 'medium'
        b = buckets.setdefault(diff, {
            'difficulty': diff,
            'label': DIFFICULTY_LABELS.get(diff, diff.title()),
            'count': 0,
            'rates_sum': 0.0,
            'rates_n': 0,
        })
        b['count'] += 1
        ac = answered_count.get(q.id, 0)
        cc = correct_count.get(q.id, 0)
        if ac > 0:
            b['rates_sum'] += (cc / ac) * 100.0
            b['rates_n'] += 1

    by_difficulty = []
    # Tartib: oddiy → qiyin.
    order = ['easy', 'medium', 'hard', 'beginner', 'elementary', 'pre-int', 'int', 'upper-int', 'advanced']
    for diff in order:
        if diff in buckets:
            b = buckets.pop(diff)
            avg_rate = round(b['rates_sum'] / b['rates_n'], 1) if b['rates_n'] else 0.0
            by_difficulty.append({
                'difficulty': b['difficulty'],
                'label': b['label'],
                'count': b['count'],
                'avg_correct_rate': avg_rate,
            })
    # Boshqa custom difficulty qiymatlari qolgan bo'lsa qo'shamiz.
    for b in buckets.values():
        avg_rate = round(b['rates_sum'] / b['rates_n'], 1) if b['rates_n'] else 0.0
        by_difficulty.append({
            'difficulty': b['difficulty'],
            'label': b['label'],
            'count': b['count'],
            'avg_correct_rate': avg_rate,
        })

    return Response({
        'total_questions': total,
        'by_difficulty': by_difficulty,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def my_monthly_stats(request):
    """GET /api/results/me/monthly/?months=6 — so'nggi N oy o'rtacha ballari.

    Profile.jsx dagi "Natijalar dinamikasi" BarChart uchun. Hardcoded
    [72,81,87,83,91] o'rniga real oylik o'rtacha qiymat.
    """
    try:
        months = int(request.query_params.get('months') or 6)
    except (TypeError, ValueError):
        months = 6
    months = max(1, min(months, 24))

    now = timezone.now()
    # Hozirgi oyning birinchi kuni
    current_first = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # Boshlang'ich oyni hisoblaymiz: months-1 oy orqaga
    year = current_first.year
    month = current_first.month - (months - 1)
    while month <= 0:
        month += 12
        year -= 1
    range_start = current_first.replace(year=year, month=month)

    # Avval har oy uchun alohida aggregate query (N+1) edi. Endi bitta
    # TruncMonth bilan oylik aggregate olib, Python dict orqali bucket'larga
    # tarqatamiz.
    raw = (TestAttempt.objects
        .filter(user=request.user, submitted_at__gte=range_start)
        .annotate(month_bucket=TruncMonth('submitted_at'))
        .values('month_bucket')
        .annotate(avg=Avg('score'), count=Count('id'))
        .order_by('month_bucket'))
    by_key = {}
    for row in raw:
        mb = row['month_bucket']
        if mb is None:
            continue
        by_key[(mb.year, mb.month)] = row

    buckets = []
    for i in range(months - 1, -1, -1):
        y = current_first.year
        m = current_first.month - i
        while m <= 0:
            m += 12
            y -= 1
        row = by_key.get((y, m))
        buckets.append({
            'year': y,
            'month': m,
            'label': f'{m}-oy',
            'average_score': round((row['avg'] if row else 0) or 0, 1),
            'attempts': (row['count'] if row else 0) or 0,
        })
    return Response({'months': buckets})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def leaderboard(request):
    """GET /api/leaderboard/?olympiad=<id>  — ranked attempts.

    Without ``olympiad`` query param, returns the top scores within the
    visible events: public olympiads globally, center competitions for the
    user's approved centers.
    """
    # Diskvalifitsiya bo'lgan attempt'lar leaderboard'da ko'rinmaydi —
    # ular faqat manager statistikasi uchun yoziladi. Aks holda 0 balli
    # ko'p qator rank'ni chalkash qilardi.
    qs = (
        TestAttempt.objects
        .filter(disqualified=False, olympiad__is_deleted=False)
        .select_related('user', 'olympiad', 'olympiad__center')
        .order_by('-score', 'time_spent', 'submitted_at')
    )
    olympiad_id = request.query_params.get('olympiad')
    if olympiad_id:
        olympiad = get_object_or_404(Olympiad.objects.select_related('center'), pk=olympiad_id)
        if not user_can_participate_in_event(request.user, olympiad):
            return Response({'detail': 'Forbidden'}, status=http_status.HTTP_403_FORBIDDEN)
        # Draft/inactive olimpiada leaderboard'i faqat manager/owner/admin
        # uchun ko'rinadi. Oddiy ishtirokchi DRAFT yoki INACTIVE tadbir
        # natijalarini ko'ra olmaydi — bu sizdirib qo'yiladigan ma'lumot.
        if olympiad.status not in (Olympiad.STATUS_ACTIVE, Olympiad.STATUS_FINISHED):
            if not user_can_manage_center_event(request.user, olympiad.center):
                return Response(
                    {'detail': 'Bu olimpiada leaderboard\'i hali ochilmagan'},
                    status=http_status.HTTP_403_FORBIDDEN,
                )
        qs = qs.filter(olympiad=olympiad)
    if not olympiad_id:
        allowed_center_ids = list(CenterMembership.objects.filter(
            user=request.user, status=CenterMembership.STATUS_APPROVED,
        ).values_list('center_id', flat=True))
        # Avval `qs.filter(...) | qs.filter(...)` orqali OR amali bajarilardi —
        # bu Django'da ikkita JOIN va katta order_by zanjirini hosil qilib,
        # query plan'ni murakkablashtirar va bir qator versiyalarda dublikat
        # qator qaytarardi. Yagona Q() bilan yaxshiroq.
        qs = qs.filter(
            Q(olympiad__event_type=Olympiad.EVENT_TYPE_OLYMPIAD)
            | Q(
                olympiad__event_type=Olympiad.EVENT_TYPE_COMPETITION,
                olympiad__center_id__in=allowed_center_ids,
            )
        )
        # Draft/inactive olimpiadalar global leaderboard'da ko'rinmasin —
        # faqat active yoki finished holatdagi tadbirlar ishtirokchilari.
        qs = qs.filter(
            olympiad__status__in=[Olympiad.STATUS_ACTIVE, Olympiad.STATUS_FINISHED]
        )
        # Global rejimda har foydalanuvchi faqat BIR marta ko'rinadi — eng
        # yaxshi attempti bilan (eng yuqori ball, teng bo'lsa eng tez vaqt).
        # Avval har bir attempt alohida qator edi va 5 ta olimpiadada
        # qatnashgan o'quvchi reytingda 5 marta chiqardi. Window RowNumber
        # (PARTITION BY user) bilan per-user birinchi qatornigina olamiz —
        # Django 4.2+ window funksiyalar bo'yicha filtrni qo'llab-quvvatlaydi.
        from django.db.models import F, Window
        from django.db.models.functions import RowNumber
        qs = qs.annotate(
            _user_row=Window(
                expression=RowNumber(),
                partition_by=[F('user_id')],
                order_by=[
                    F('score').desc(),
                    F('time_spent').asc(),
                    F('submitted_at').asc(),
                ],
            )
        ).filter(_user_row=1)
        # Window filtri queryset'ni subquery'ga o'raydi — yakuniy tartibni
        # qayta tiklaymiz (eng yuqori ball yuqorida).
        qs = qs.order_by('-score', 'time_spent', 'submitted_at')
    # Pagination: `?page=` va `?page_size=` query parametrlari qo'llab-
    # quvvatlanadi. Default page_size=100, maksimum 500. Eski `?limit=`
    # parametri ham backward-compat uchun qabul qilinadi.
    try:
        page = int(request.query_params.get('page') or 1)
    except (TypeError, ValueError):
        page = 1
    page = max(1, page)
    try:
        page_size = int(
            request.query_params.get('page_size')
            or request.query_params.get('limit')
            or 100
        )
    except (TypeError, ValueError):
        page_size = 100
    page_size = max(1, min(page_size, 500))
    total_count = qs.count()
    offset = (page - 1) * page_size
    qs = qs[offset:offset + page_size]
    # Rank submit ichida yangilanmaydi (DB yukini kamaytirish uchun). Shu
    # sababli leaderboard'da har doim joriy tartiblash (`-score`,
    # `time_spent`, `submitted_at`) bo'yicha `i+1` o'rin beriladi. Bu
    # filter (masalan, faqat bitta olimpiada) uchun ham to'g'ri natija
    # qaytaradi, chunki tartiblash querysetda allaqachon qo'llanilgan.
    from accounts.utils import avatar_url_for
    entries = []
    for i, a in enumerate(qs):
        # Public olimpiadalarda `center` NULL bo'lishi mumkin — `a.olympiad.
        # center.name` to'g'ridan-to'g'ri o'qilsa AttributeError (500) berardi.
        # Markaz bo'lmasa markazga bog'liq maydonlar bo'sh qaytariladi.
        center = a.olympiad.center if a.olympiad.center_id else None
        entries.append({
            'rank': offset + i + 1,
            'attempt_id': a.id,
            'user_id': a.user_id,
            'name': a.user.full_name,
            'avatar_url': avatar_url_for(a.user, request),
            'is_premium': a.user.is_premium,
            'center': center.name if center else '',
            'organization_type': center.organization_type if center else '',
            'country': center.country if center else '',
            'region': center.region if center else '',
            'district': center.district if center else '',
            'subject': a.olympiad.subject,
            'olympiad_id': a.olympiad_id,
            'olympiad_title': a.olympiad.title,
            'olympiad_status': a.olympiad.status,
            'score': a.score,
            'time_spent': a.time_spent,
            'submitted_at': a.submitted_at.isoformat(),
        })
    pagination_meta = {
        'page': page,
        'page_size': page_size,
        'total': total_count,
        'has_next': offset + len(entries) < total_count,
    }
    # Header info: tanlangan olympiad bo'lsa olympiad ma'lumoti, aks holda
    # eng ko'p kelgan olympiad nomi (frontend subtitle uchun).
    header = None
    if olympiad_id:
        olym = Olympiad.objects.select_related('center').filter(pk=olympiad_id).first()
        if olym:
            header = {
                'olympiad_id': olym.id,
                'olympiad_title': olym.title,
                'subject': olym.subject,
                'status': olym.status,
                'start_datetime': olym.start_datetime.isoformat() if olym.start_datetime else None,
            }
    elif entries:
        # Eng yaqinda topshirilgan attemptdan olympiad nomini olamiz
        latest = entries[0]
        header = {
            'olympiad_id': latest.get('olympiad_id'),
            'olympiad_title': latest.get('olympiad_title'),
            'subject': latest.get('subject'),
            'status': latest.get('olympiad_status'),
            'start_datetime': latest.get('submitted_at'),
        }

    return Response({
        'results': entries,
        'pagination': pagination_meta,
        'header': header,
    })



@api_view(['POST'])
@permission_classes([IsAuthenticated])
@throttle_classes([ScopedRateThrottle])
def test_session_ping(request):
    """POST /api/attempts/ping/

    Body: {"olympiad": <id>, "answered_count": <int>, "tab_escapes": <int>,
           "device_id": <str>}
    o'quvchining joriy holatini cache'da yangilab borish uchun ping.

    Parallel sessiya tekshiruvi: agar oxirgi ping boshqa device_id'dan kelgan
    bo'lsa va orada 30 soniyadan kam vaqt o'tgan bo'lsa (ya'ni ikkala qurilma
    ham faol) — session DQ qilinadi va 409 qaytariladi.
    """
    from datetime import timedelta

    from django.core.cache import cache
    from django.utils import timezone

    data = request.data or {}
    try:
        olympiad_id = int(data.get('olympiad'))
    except (TypeError, ValueError):
        return Response({'detail': "olympiad majburiy"}, status=http_status.HTTP_400_BAD_REQUEST)

    try:
        answered_count = int(data.get('answered_count', 0))
        tab_escapes = int(data.get('tab_escapes', 0))
    except (TypeError, ValueError):
        answered_count = 0
        tab_escapes = 0

    # device_id body yoki header'dan olinadi (frontend yuboradi).
    device_id = str(
        data.get('device_id')
        or request.META.get('HTTP_X_DEVICE_ID')
        or ''
    )[:64]

    now = timezone.now()

    # Race condition himoyasi: ikkita ping (masalan, ikki qurilmadan) bir
    # vaqtda kelib bir-birining `last_device_id`/`last_ping_at` qiymatini
    # ustiga yozib yuborishi mumkin edi — natijada parallel sessiya tekshiruvi
    # ishlamay qolardi. Sessiyani `select_for_update()` bilan qatorni lock
    # qilib, tekshiruv + yangilashni atomic blok ichida bajaramiz.
    concurrent_conflict = False
    with transaction.atomic():
        session = (
            TestSession.objects
            .select_for_update()
            .filter(user=request.user, olympiad_id=olympiad_id)
            .first()
        )
        if not session:
            return Response({'detail': "Test sessiya topilmadi"}, status=http_status.HTTP_404_NOT_FOUND)

        # Faqat ACTIVE sessiyada parallel-qurilma tekshiruvi qilinadi.
        # COMPLETED sessiyaga kechikib kelgan ping (masalan, submit'dan keyin
        # navbatda qolgan so'rov) foydalanuvchini nohaq 409/DQ qilmasin.
        # DISQUALIFIED holat esa quyidagi oqimda 409 qaytarishda davom etadi —
        # frontend shu javob orqali cheat ekranini ko'rsatadi.
        if session.status == TestSession.STATUS_COMPLETED:
            return Response({'ok': True, 'status': session.status})

        # Parallel sessiya tekshiruvi — device_id berilgan bo'lsa.
        if device_id:
            if (
                session.last_device_id
                and session.last_device_id != device_id
                and session.last_ping_at
                and (now - session.last_ping_at) < timedelta(seconds=30)
            ):
                # Boshqa qurilma 30 soniya ichida faol — bir vaqtda ikki kirish.
                if session.status != TestSession.STATUS_DISQUALIFIED:
                    session.status = TestSession.STATUS_DISQUALIFIED
                    session.disqualified_at = session.disqualified_at or now
                    session.cheating_reason = session.cheating_reason or 'concurrent_session'
                    session.save(update_fields=['status', 'disqualified_at', 'cheating_reason'])
                concurrent_conflict = True
            else:
                # Aks holda joriy qurilmani egasi deb belgilaymiz.
                session.last_device_id = device_id
                session.last_ping_at = now
                session.save(update_fields=['last_device_id', 'last_ping_at'])

    if concurrent_conflict:
        return Response(
            {
                'disqualified': True,
                'detail': "Boshqa qurilmadan kirilgani aniqlandi. Olimpiada yakunlandi.",
            },
            status=http_status.HTTP_409_CONFLICT,
        )

    # Cacheni yangilash (timeout 60 soniya, agar 60s ichida ping kelmasa oflayn hisoblanadi)
    cache_key = f"test_session_ping:{olympiad_id}:{request.user.id}"
    cache.set(cache_key, {
        'answered_count': answered_count,
        'tab_escapes': tab_escapes,
        'last_ping': now.isoformat(),
    }, timeout=60)

    return Response({'ok': True})


test_session_ping.cls.throttle_scope = 'ping'


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def olympiad_live_proctoring(request, olympiad_id):
    """GET /api/manager/olympiads/<id>/live/

    Manager uchun faol o'quvchilar ro'yxati va real vaqtda holatlari.
    """
    from django.core.cache import cache
    from django.utils import timezone
    from datetime import datetime

    olympiad = get_object_or_404(Olympiad.objects.select_related('center'), pk=olympiad_id)
    if not _user_can_manage_olympiad(request.user, olympiad):
        return Response({'detail': 'Forbidden'}, status=http_status.HTTP_403_FORBIDDEN)

    sessions = list(
        TestSession.objects
        .filter(olympiad=olympiad)
        .select_related('user')
        .order_by('-started_at')
    )

    attempts = {
        a.user_id: a
        for a in TestAttempt.objects.filter(olympiad=olympiad)
    }

    # Loop ichida har iteratsiyada count() qilmaslik uchun bir marta hisoblaymiz
    olympiad_question_count = olympiad.questions.count()

    # Ping ma'lumotlarini bitta cache.get_many() chaqiruvi bilan olamiz —
    # avval loop ichida har o'quvchi uchun alohida cache.get() chaqirilardi
    # (N ta Redis round-trip). Endi barcha kalitlarni oldindan list qilib,
    # bitta so'rovda hammasini olamiz.
    ping_keys = {
        s.user_id: f"test_session_ping:{olympiad_id}:{s.user_id}"
        for s in sessions
    }
    ping_map = cache.get_many(list(ping_keys.values()))

    from accounts.utils import avatar_url_for
    now = timezone.now()
    results = []

    for s in sessions:
        user = s.user
        attempt = attempts.get(user.id)

        # Cache'dan joriy ping ma'lumotlarini o'qish (get_many natijasidan)
        ping_data = ping_map.get(ping_keys.get(user.id))

        # O'quvchi onlaynmi yoki yo'qligini tekshirish
        is_online = False
        answered = 0
        escapes = 0

        if ping_data:
            last_ping_str = ping_data.get('last_ping')
            try:
                last_ping = datetime.fromisoformat(last_ping_str)
                # 45 soniyadan kam vaqt ichida ping kelgan bo'lsa onlayn deb olamiz
                if (now - last_ping).total_seconds() < 45:
                    is_online = True
            except Exception:
                pass
            answered = ping_data.get('answered_count', 0)
            escapes = ping_data.get('tab_escapes', 0)

        status = 'active'
        if attempt:
            status = 'disqualified' if attempt.disqualified else 'completed'
            is_online = False
            answered = attempt.total_questions
        elif s.status == TestSession.STATUS_DISQUALIFIED:
            status = 'disqualified'
            is_online = False

        results.append({
            'student_id': user.id,
            'student_name': user.full_name or user.phone or 'O\'quvchi',
            'avatar_url': avatar_url_for(user, request),
            'phone': user.normalized_phone or user.phone or '—',
            'started_at': s.started_at.isoformat(),
            'status': status,
            'cheating_reason': s.cheating_reason,
            'answered_count': answered,
            'total_questions': len(s.question_order) or olympiad_question_count or 0,
            'tab_escapes': escapes,
            'is_online': is_online,
            'score': attempt.score if attempt else None,
            'time_spent': attempt.time_spent if attempt else None,
        })

    return Response(results)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@throttle_classes([ScopedRateThrottle])
def get_mistakes_list(request):
    """GET /api/attempts/mistakes/
    O'quvchining barcha o'tgan imtihonlardagi noto'g'ri berilgan javoblarini savollar kesimida yig'ib qaytaradi.
    """
    from questions.models import Question
    from .models import TestAttempt

    # Faqat `answers` ustunini stream qilamiz (butun TestAttempt qatorlarini
    # xotiraga yuklamaymiz). `.iterator()` server-side kursor bilan ko'p
    # attempt'li foydalanuvchilarda xotirani bir tekis ushlab turadi.
    answers_stream = (
        TestAttempt.objects
        .filter(user=request.user, disqualified=False)
        .values_list('answers', flat=True)
        .iterator()
    )

    # Avval har bir savol uchun (birinchi ko'rilgan) tanlangan javobni yig'amiz,
    # keyin barcha savollarni bitta so'rov bilan olamiz (N+1'ni oldini olish uchun).
    # chosen_val — savol turiga qarab int/str/list/dict bo'lishi mumkin.
    # int() ga majburlamaymiz: aks holda fill_blank/multiple_select/essay
    # javoblari ValueError bilan tushib qolib, xatolar ro'yxatiga umuman
    # kirmasdi. Xom qiymatni saqlaymiz va grade_answer bilan baholaymiz.
    chosen_by_question = {}
    for answers in answers_stream:
        answers = answers or {}
        for q_id_str, chosen_val in answers.items():
            try:
                q_id = int(q_id_str)
            except (ValueError, TypeError):
                continue
            if q_id in chosen_by_question:
                continue
            chosen_by_question[q_id] = chosen_val

    questions_by_id = {
        q.id: q
        for q in Question.objects.filter(pk__in=chosen_by_question.keys())
    }

    mistakes = []
    for q_id, chosen_val in chosen_by_question.items():
        question = questions_by_id.get(q_id)
        if not question:
            continue
        # Faqat aniq noto'g'ri javoblar (RESULT_WRONG) xatolar ro'yxatiga
        # kiradi: to'g'ri (correct), bo'sh (blank) va qo'lda baholanadigan
        # essay (pending_review) chiqarib tashlanadi.
        if grade_answer(question, chosen_val) == RESULT_WRONG:
            mistakes.append({
                'question_id': question.id,
                'subject': question.subject,
                'text': question.text,
                'options': question.options,
                'correct_answer': question.correct_answer,
                'chosen_answer': chosen_val,
                'explanation': question.explanation or '',
            })
    return Response(mistakes)


get_mistakes_list.cls.throttle_scope = 'mistakes'


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@throttle_classes([ScopedRateThrottle])
def explain_all_mistakes(request):
    """POST /api/attempts/mistakes/explain/
    O'quvchining xatolari asosida Gemini AI orqali umumiy tavsiyalar generatsiya qiladi.
    """
    from questions.models import Question
    from questions.ai_generation import explain_mistakes_ai
    from .models import TestAttempt

    # Avval har bir savol uchun (birinchi ko'rilgan) tanlangan javobni yig'amiz,
    # tartibni saqlagan holda (dict insertion order). Keyin barcha savollarni
    # bitta so'rov bilan olamiz (N+1'ni oldini olish uchun). Faqat `answers`
    # JSON maydonini .values_list(...).iterator() bilan oqimda o'qiymiz —
    # to'liq TestAttempt obyektlarini xotiraga yuklab, ko'p attempt'li
    # foydalanuvchida xotirani to'ldirib qo'ymaslik uchun (xuddi
    # get_mistakes_list dagi kabi).
    answers_iter = (
        TestAttempt.objects
        .filter(user=request.user, disqualified=False)
        .values_list('answers', flat=True)
        .iterator()
    )
    # Xom javob qiymatini saqlaymiz (int() ga majburlamaymiz) — yangi savol
    # turlari (fill_blank/multiple_select/essay) javoblari ham yig'ilsin.
    chosen_by_question = {}
    for answers in answers_iter:
        answers = answers or {}
        for q_id_str, chosen_val in answers.items():
            try:
                q_id = int(q_id_str)
            except (ValueError, TypeError):
                continue
            if q_id in chosen_by_question:
                continue
            chosen_by_question[q_id] = chosen_val

    questions_by_id = {
        q.id: q
        for q in Question.objects.filter(pk__in=chosen_by_question.keys())
    }

    mistakes = []
    for q_id, chosen_val in chosen_by_question.items():
        question = questions_by_id.get(q_id)
        if not question:
            continue
        # Faqat aniq noto'g'ri javoblarni (RESULT_WRONG) AI tahliliga beramiz.
        if grade_answer(question, chosen_val) == RESULT_WRONG:
            mistakes.append({
                'question_id': question.id,
                'subject': question.subject,
                'text': question.text,
                'options': question.options,
                'correct_answer': question.correct_answer,
                'chosen_answer': chosen_val,
            })
            if len(mistakes) >= 8:
                break

    if not mistakes:
        return Response({'explanation': "Sizda hozircha xatolar aniqlanmadi. Barakalla!"})

    explanation_text = explain_mistakes_ai(mistakes)

    # T5: AI xatolar tahlilini menejer faoliyat logiga yozamiz (markaz bo'lsa).
    try:
        from centers.models import ManagerActivityLog
        from centers.services import log_manager_activity, primary_center_for_user
        center = primary_center_for_user(request.user)
        if center is not None:
            log_manager_activity(
                center, request.user, ManagerActivityLog.ACTION_SEND_ANALYSIS,
                description='Xatolar tahlili (AI) generatsiya qilindi',
                target_user=request.user,
            )
    except Exception:
        import logging
        logging.getLogger(__name__).exception(
            'manager activity log failed for user=%s', request.user.pk,
        )

    return Response({'explanation': explanation_text})


explain_all_mistakes.cls.throttle_scope = 'ai'


