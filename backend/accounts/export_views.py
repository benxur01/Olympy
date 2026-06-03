"""Excel eksport endpointlari.

Markaz a'zolari va olimpiada natijalarini .xlsx formatida yuklab olish.
openpyxl ishlatiladi (requirements.txt'da pinlangan). Maydon nomlari haqiqiy
modellarga moslashtirilgan:
  - TestAttempt vaqt maydoni: `time_spent` (soniya), sana: `submitted_at`
  - CenterMembership: `created_at`
  - Olympiad: `title`
"""
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO


def _style_header(ws, headers):
    """Sarlavha qatorini ko'k fon + oq qalin shrift bilan bezaydi."""
    header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')


def _xlsx_response(wb, filename):
    """Workbook'ni HttpResponse (attachment) sifatida qaytaradi."""
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_center_members_excel(request, center_id):
    """Markaz a'zolari ro'yxatini Excel formatda yuklab olish."""
    from centers.models import EducationCenter, CenterMembership

    center = EducationCenter.objects.filter(pk=center_id).first()
    if not center:
        return Response({'detail': 'Markaz topilmadi'}, status=404)

    # Permission tekshiruv — owner/admin/manager yoki platforma admini.
    membership = CenterMembership.objects.filter(
        center=center, user=request.user,
        role__in=['owner', 'admin', 'manager'], status='approved',
    ).first()
    if not membership and not request.user.is_staff:
        return Response({'detail': 'Ruxsat yo\'q'}, status=403)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "A'zolar"

    headers = ['#', 'To\'liq ism', 'Telefon', 'Rol', 'Status', 'Qo\'shilgan']
    _style_header(ws, headers)

    members = (
        CenterMembership.objects
        .filter(center=center, status='approved')
        .select_related('user')
        .order_by('-created_at')
    )
    for row, m in enumerate(members, 2):
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=m.user.full_name)
        ws.cell(row=row, column=3, value=m.user.phone or '')
        ws.cell(row=row, column=4, value=m.role)
        ws.cell(row=row, column=5, value=m.status)
        ws.cell(row=row, column=6, value=m.created_at.strftime('%Y-%m-%d') if m.created_at else '')

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    return _xlsx_response(wb, f'members_{center_id}.xlsx')


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_olympiad_results_excel(request, olympiad_id):
    """Olimpiada natijalarini Excel formatda yuklab olish."""
    from olympiads.models import Olympiad
    from attempts.models import TestAttempt

    olympiad = Olympiad.objects.filter(pk=olympiad_id).first()
    if not olympiad:
        return Response({'detail': 'Olimpiada topilmadi'}, status=404)

    # Permission tekshiruv — olimpiada egasi markazning owner/admin/manager'i
    # yoki platforma admini. Markazsiz (public) olimpiadalar uchun faqat
    # platforma admini eksport qila oladi.
    is_allowed = request.user.is_staff
    if not is_allowed and olympiad.center_id:
        from centers.models import CenterMembership
        is_allowed = CenterMembership.objects.filter(
            center_id=olympiad.center_id, user=request.user,
            role__in=['owner', 'admin', 'manager'], status='approved',
        ).exists()
    if not is_allowed:
        return Response({'detail': 'Ruxsat yo\'q'}, status=403)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Natijalar"

    headers = ['#', 'Ism', 'Telefon', 'Ball', 'To\'g\'ri', 'Noto\'g\'ri', 'Vaqt (min)', 'Sana']
    _style_header(ws, headers)

    attempts = (
        TestAttempt.objects
        .filter(olympiad=olympiad, disqualified=False)
        .select_related('user')
        .order_by('-score')
    )
    for row, a in enumerate(attempts, 2):
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=a.user.full_name)
        ws.cell(row=row, column=3, value=a.user.phone or '')
        ws.cell(row=row, column=4, value=float(a.score) if a.score else 0)
        ws.cell(row=row, column=5, value=a.correct_count)
        ws.cell(row=row, column=6, value=a.wrong_count)
        # TestAttempt'da vaqt maydoni `time_spent` (soniya) deb nomlangan.
        duration = a.time_spent
        ws.cell(row=row, column=7, value=round(duration / 60, 1) if duration else '')
        ws.cell(row=row, column=8, value=a.submitted_at.strftime('%Y-%m-%d %H:%M') if a.submitted_at else '')

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    return _xlsx_response(wb, f'results_{olympiad_id}.xlsx')
